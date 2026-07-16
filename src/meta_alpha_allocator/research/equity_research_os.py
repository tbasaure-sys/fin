from __future__ import annotations

import base64
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import json
import math
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from ..data.fmp_client import FMPClient
from ..data.sec_edgar_client import SECEdgarClient
from .equity_research_agents import build_agent_outputs, run_final_orchestrator_llm
from .institutional_valuation import build_institutional_valuation


DEFAULT_TAX_RATE = 0.21
DEFAULT_WACC = 0.09
DEFAULT_TERMINAL_GROWTH = 0.03
CURRENT_SHARE_COUNT_MAXIMUM_RELATIVE_DIFFERENCE = 0.20
CURRENT_SHARE_COUNT_MAXIMUM_AGE_DAYS = 45
_MODEL_XLSX_HEADLINE_FORMULA_STATUS = {
    "range_low": {"formula_driven": False, "reconciliation_difference": None},
    "range_central": {"formula_driven": False, "reconciliation_difference": None},
    "range_high": {"formula_driven": False, "reconciliation_difference": None},
    "selected_value": {"formula_driven": False, "reconciliation_difference": None},
    "bear_value_per_share": {"formula_driven": False, "reconciliation_difference": None},
    "base_value_per_share": {"formula_driven": False, "reconciliation_difference": None},
    "bull_value_per_share": {"formula_driven": False, "reconciliation_difference": None},
}
_MODEL_XLSX_RECONCILIATION_TOLERANCE = 1e-9


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_ticker(value: str) -> str:
    ticker = "".join(ch for ch in str(value or "").upper().strip() if ch.isalnum() or ch in {".", "-"})
    return ticker[:16]


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if not value:
            return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def _model_xlsx_formula_contract_is_ready() -> bool:
    """Fail closed until every decision output is formula-driven and ties to zero."""
    for status in _MODEL_XLSX_HEADLINE_FORMULA_STATUS.values():
        difference = _safe_float(status.get("reconciliation_difference"))
        if status.get("formula_driven") is not True:
            return False
        if difference is None or abs(difference) > _MODEL_XLSX_RECONCILIATION_TOLERANCE:
            return False
    return True


def _ratio(numerator: Any, denominator: Any) -> float | None:
    n = _safe_float(numerator)
    d = _safe_float(denominator)
    if n is None or d in (None, 0):
        return None
    result = n / d
    return result if math.isfinite(result) else None


def _fmt_currency(value: float | None) -> str:
    if value is None:
        return "n/a"
    abs_value = abs(value)
    if abs_value >= 1_000_000_000:
        return f"${value / 1_000_000_000:,.1f}B"
    if abs_value >= 1_000_000:
        return f"${value / 1_000_000:,.1f}M"
    return f"${value:,.0f}"


def _fmt_pct(value: float | None) -> str:
    return "n/a" if value is None else f"{value * 100:.1f}%"


def _first_existing(row: pd.Series, names: list[str]) -> Any:
    for name in names:
        if name in row and row.get(name) is not None:
            value = row.get(name)
            try:
                if pd.isna(value):
                    continue
            except TypeError:
                pass
            return value
    return None


def _sum_existing(row: pd.Series, names: list[str]) -> float | None:
    values = [_safe_float(_first_existing(row, [name])) for name in names if name in row]
    values = [value for value in values if value is not None]
    return sum(values) if values else None


def calculate_revenue_cagr(start_revenue: float, end_revenue: float, years: int) -> float | None:
    if years <= 0 or start_revenue <= 0 or end_revenue <= 0:
        return None
    return (end_revenue / start_revenue) ** (1 / years) - 1


@dataclass(frozen=True)
class DcfScenarioInput:
    latest_revenue: float
    latest_fcf_margin: float
    cash: float
    debt: float
    diluted_shares: float
    revenue_growth: float
    terminal_fcf_margin: float
    wacc: float
    terminal_growth: float
    years: int = 5


def build_dcf_scenario(name: str, payload: DcfScenarioInput) -> dict[str, Any]:
    if payload.diluted_shares <= 0:
        raise ValueError("diluted_shares must be positive")
    if payload.latest_revenue <= 0:
        raise ValueError("latest_revenue must be positive")
    if payload.wacc <= payload.terminal_growth:
        raise ValueError("wacc must be greater than terminal growth")

    forecast_rows: list[dict[str, Any]] = []
    revenue = payload.latest_revenue
    pv_fcff = 0.0
    for year in range(1, payload.years + 1):
        revenue *= 1 + payload.revenue_growth
        fcff = revenue * payload.terminal_fcf_margin
        discount_factor = (1 + payload.wacc) ** year
        present_value = fcff / discount_factor
        pv_fcff += present_value
        forecast_rows.append(
            {
                "year": year,
                "revenue": revenue,
                "free_cash_flow": fcff,
                "discount_factor": discount_factor,
                "present_value": present_value,
            }
        )

    terminal_fcff = forecast_rows[-1]["free_cash_flow"] * (1 + payload.terminal_growth)
    terminal_value = terminal_fcff / (payload.wacc - payload.terminal_growth)
    pv_terminal_value = terminal_value / ((1 + payload.wacc) ** payload.years)
    enterprise_value = pv_fcff + pv_terminal_value
    equity_value = enterprise_value + payload.cash - payload.debt
    intrinsic_value_per_share = equity_value / payload.diluted_shares

    return {
        "name": name,
        "assumptions": {
            "revenue_growth": payload.revenue_growth,
            "terminal_fcf_margin": payload.terminal_fcf_margin,
            "wacc": payload.wacc,
            "terminal_growth": payload.terminal_growth,
            "years": payload.years,
        },
        "forecast": forecast_rows,
        "pv_fcff": pv_fcff,
        "pv_terminal_value": pv_terminal_value,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "intrinsic_value_per_share": intrinsic_value_per_share,
    }


def reverse_dcf_implied_growth(
    *,
    current_price: float,
    latest_revenue: float,
    fcf_margin: float,
    cash: float,
    debt: float,
    diluted_shares: float,
    wacc: float = DEFAULT_WACC,
    terminal_growth: float = DEFAULT_TERMINAL_GROWTH,
    years: int = 5,
) -> dict[str, Any]:
    if current_price <= 0 or latest_revenue <= 0 or diluted_shares <= 0 or fcf_margin <= 0:
        return {"available": False, "reason": "missing current price, revenue, shares, or positive FCF margin"}

    def value_at(growth: float) -> float:
        scenario = build_dcf_scenario(
            "reverse",
            DcfScenarioInput(
                latest_revenue=latest_revenue,
                latest_fcf_margin=fcf_margin,
                cash=cash,
                debt=debt,
                diluted_shares=diluted_shares,
                revenue_growth=growth,
                terminal_fcf_margin=fcf_margin,
                wacc=wacc,
                terminal_growth=terminal_growth,
                years=years,
            ),
        )
        return float(scenario["intrinsic_value_per_share"])

    low = -0.15
    high = 0.35
    low_value = value_at(low)
    high_value = value_at(high)
    status = "solved"
    if current_price < low_value:
        status = "below_range"
        solved = low
    elif current_price > high_value:
        status = "above_range"
        solved = high
    else:
        solved = 0.0
        for _ in range(80):
            mid = (low + high) / 2
            mid_value = value_at(mid)
            if mid_value < current_price:
                low = mid
            else:
                high = mid
            solved = (low + high) / 2

    return {
        "available": True,
        "status": status,
        "implied_revenue_cagr": solved,
        "current_price": current_price,
        "value_at_floor": low_value,
        "value_at_ceiling": high_value,
        "assumptions": {
            "terminal_fcf_margin": fcf_margin,
            "wacc": wacc,
            "terminal_growth": terminal_growth,
            "years": years,
        },
    }


def _source_record(source_id: str, provider: str, endpoint: str, status: str, **extra: Any) -> dict[str, Any]:
    if extra.get("error") is not None:
        extra["error"] = _redact_secret_text(str(extra["error"]))
    return {
        "source_id": source_id,
        "provider": provider,
        "endpoint_or_filing": endpoint,
        "retrieved_at": _now_iso(),
        "status": status,
        **extra,
    }


def _redact_secret_text(text: str) -> str:
    redacted = re.sub(r"(?i)(apikey=)[^&\s]+", r"\1[redacted]", text)
    redacted = re.sub(r"(?i)(api[_-]?key['\"]?\s*[:=]\s*['\"]?)[^,'\"\s}]+", r"\1[redacted]", redacted)
    redacted = re.sub(r"(?i)(authorization:\s*bearer\s+)[^,\s}]+", r"\1[redacted]", redacted)
    return redacted


def _data_point(metric: str, value: Any, tag: str, source_id: str | None = None, formula: str | None = None) -> dict[str, Any]:
    return {
        "metric": metric,
        "raw_value": value,
        "normalized_value": value,
        "claim_tag": tag,
        "source_id": source_id,
        "formula": formula,
    }


EXPECTED_EVIDENCE_METRICS = [
    "company_profile",
    "latest_revenue",
    "latest_diluted_shares",
    "latest_free_cash_flow",
    "revenue_cagr_5y",
    "gross_margin",
    "operating_margin",
    "fcf_margin",
    "roic",
    "net_debt",
    "base_fcf_margin",
    "wacc",
    "terminal_growth",
    "current_price",
    "valuation_range_central",
    "reverse_dcf_status",
    "ev_to_sales",
    "price_to_fcf",
    "latest_sec_filing",
]

DEFAULT_STATEMENT_SOURCE_IDS = {
    "income": "fmp:income:annual",
    "cash_flow": "fmp:cash-flow:annual",
    "balance": "fmp:balance:annual",
}

SEC_STATEMENT_SOURCE_IDS = {
    "income": "sec:companyfacts:income",
    "cash_flow": "sec:companyfacts:cash-flow",
    "balance": "sec:companyfacts:balance",
}
QUARTERLY_STATEMENT_SOURCE_IDS = {
    "income": "fmp:income:quarterly",
    "cash_flow": "fmp:cash-flow:quarterly",
    "balance": "fmp:balance:quarterly",
}
TTM_STATEMENT_SOURCE_IDS = {
    "income": "fmp:income:ttm",
    "cash_flow": "fmp:cash-flow:ttm",
    "balance": "fmp:balance:ttm",
}
YFINANCE_STATEMENT_SOURCE_IDS = {
    "income": "yfinance:income:annual",
    "cash_flow": "yfinance:cash-flow:annual",
    "balance": "yfinance:balance:annual",
}
YFINANCE_QUARTERLY_STATEMENT_SOURCE_IDS = {
    "income": "yfinance:income:quarterly",
    "cash_flow": "yfinance:cash-flow:quarterly",
    "balance": "yfinance:balance:quarterly",
}
SEC_RECONCILIATION_SOURCE_ID = "sec:companyfacts:reconciliation"

SEC_ANNUAL_FORMS = {"10-K", "10-K/A", "20-F", "20-F/A", "40-F", "40-F/A"}

SEC_CONCEPTS = {
    "income": {
        "revenue": ["RevenueFromContractWithCustomerExcludingAssessedTax", "Revenues", "SalesRevenueNet"],
        "grossProfit": ["GrossProfit"],
        "costOfRevenue": ["CostOfRevenue", "CostOfGoodsAndServicesSold"],
        "operatingIncome": ["OperatingIncomeLoss"],
        "incomeBeforeTax": [
            "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
            "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments",
        ],
        "incomeTaxExpense": ["IncomeTaxExpenseBenefit"],
        "netIncome": ["NetIncomeLoss"],
        "interestExpense": ["InterestExpenseNonOperating", "InterestAndDebtExpense"],
        "weightedAverageShsOutDil": ["WeightedAverageNumberOfDilutedSharesOutstanding"],
    },
    "cash_flow": {
        "netCashProvidedByOperatingActivities": [
            "NetCashProvidedByUsedInOperatingActivities",
            "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations",
        ],
        "capitalExpenditure": ["PaymentsToAcquirePropertyPlantAndEquipment", "PaymentsToAcquireProductiveAssets"],
        "depreciationAndAmortization": ["DepreciationDepletionAndAmortization", "DepreciationDepletionAndAmortizationExpense"],
        "stockBasedCompensation": ["ShareBasedCompensation"],
        "commonStockRepurchased": ["PaymentsForRepurchaseOfCommonStock", "PaymentsForRepurchaseOfEquity"],
    },
    "balance": {
        "cashAndCashEquivalents": ["CashAndCashEquivalentsAtCarryingValue", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"],
        "totalDebt": ["LongTermDebtAndFinanceLeaseObligations", "LongTermDebt"],
        "shortTermDebt": ["LongTermDebtAndFinanceLeaseObligationsCurrent", "LongTermDebtCurrent", "ShortTermBorrowings", "ShortTermDebt"],
        "longTermDebt": ["LongTermDebtAndFinanceLeaseObligationsNoncurrent", "LongTermDebtNoncurrent"],
        "totalStockholdersEquity": ["StockholdersEquity", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],
        "totalAssets": ["Assets"],
        "netReceivables": ["AccountsReceivableNetCurrent", "ReceivablesNetCurrent"],
        "inventory": ["InventoryNet"],
        "goodwillAndIntangibleAssets": ["GoodwillAndIntangibleAssetsNet"],
        "goodwill": ["Goodwill"],
        "intangibleAssets": ["IntangibleAssetsNetExcludingGoodwill", "FiniteLivedIntangibleAssetsNet"],
        "pensionBenefitObligation": ["DefinedBenefitPlanBenefitObligation"],
        "pensionPlanAssets": ["DefinedBenefitPlanFairValueOfPlanAssets"],
    },
}


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _financial_source_for_field(field: str, source_ids: dict[str, str] | None = None) -> str | None:
    source_ids = source_ids or DEFAULT_STATEMENT_SOURCE_IDS
    income_fields = {
        "revenue",
        "gross_profit",
        "cost_of_revenue",
        "operating_income",
        "pretax_income",
        "tax_expense",
        "net_income",
        "ebitda",
        "diluted_shares",
    }
    cash_flow_fields = {
        "cash_from_operations",
        "capital_expenditures",
        "depreciation_amortization",
        "stock_based_compensation",
        "common_stock_repurchased",
    }
    balance_fields = {
        "cash",
        "total_debt",
        "short_term_debt",
        "long_term_debt",
        "total_equity",
        "total_assets",
        "net_receivables",
        "inventory",
        "goodwill_and_intangibles",
        "preferred_stock",
        "minority_interest",
        "unfunded_pension_liability",
        "lease_liabilities_not_in_debt",
        "non_operating_investments",
    }
    if field in income_fields:
        return source_ids.get("income")
    if field in cash_flow_fields:
        return source_ids.get("cash_flow")
    if field in balance_fields:
        return source_ids.get("balance")
    return None


def _financial_formula_for_field(field: str) -> str | None:
    formulas = {
        "free_cash_flow": "cash_from_operations - abs(capital_expenditures)",
        "fcff": "free_cash_flow + after_tax_interest_expense",
        "fcff_after_sbc": "fcff - stock_based_compensation; SBC treated as an owner cost because future dilution is not modeled",
        "nopat": "operating_income * (1 - normalized_tax_rate)",
        "invested_capital": "total_debt + total_equity - cash",
        "gross_margin": "gross_profit / revenue",
        "operating_margin": "operating_income / revenue",
        "net_margin": "net_income / revenue",
        "fcf_margin": "free_cash_flow / revenue",
        "cash_conversion": "cash_from_operations / net_income",
        "sbc_as_pct_revenue": "stock_based_compensation / revenue",
        "sbc_as_pct_fcf": "stock_based_compensation / free_cash_flow",
        "roic": "NOPAT / average_invested_capital",
    }
    return formulas.get(field)


def _financial_data_points(rows: list[dict[str, Any]], source_ids: dict[str, str] | None = None) -> list[dict[str, Any]]:
    ledger: list[dict[str, Any]] = []
    for row in rows:
        period = row.get("fiscal_year") or row.get("date") or "unknown"
        for field, value in row.items():
            if field in {"date", "fiscal_year"} or not _has_value(value):
                continue
            source_id = _financial_source_for_field(field, source_ids)
            if field == "unfunded_pension_liability" and row.get("unfunded_pension_liability_source_id"):
                source_id = str(row["unfunded_pension_liability_source_id"])
            formula = _financial_formula_for_field(field)
            if source_id:
                ledger.append(_data_point(f"financials.annual.{period}.{field}", value, "sourced_fact", source_id))
            elif formula:
                ledger.append(_data_point(f"financials.annual.{period}.{field}", value, "calculated_metric", formula=formula))
    return ledger


def _ttm_data_points(
    ttm_row: dict[str, Any] | None,
    quarterly_source_ids: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    if not ttm_row:
        return []
    quarterly_source_ids = quarterly_source_ids or QUARTERLY_STATEMENT_SOURCE_IDS
    validation = ttm_row.get("ttm_validation") or {}
    quarter_dates = validation.get("quarter_dates") or []
    provider_reconciled = validation.get("provider_ttm_reconciled") is True
    provider_balance_reconciled = bool(
        validation.get("provider_ttm_balance_date_current") is True
        and validation.get("provider_ttm_balance_currency_reconciled") is True
    )
    provider_reconciled_metrics = set(validation.get("provider_reconciled_metrics") or [])
    sec_reconciled_metrics = set(validation.get("sec_reconciled_metrics") or [])

    def metric_is_supported(field: str, reconciled_metrics: set[str]) -> bool:
        direct_dependencies = {
            "free_cash_flow": {"cash_from_operations", "capital_expenditures"},
            "fcff": {"cash_from_operations", "capital_expenditures", "interest_expense"},
            "fcff_after_sbc": {
                "cash_from_operations",
                "capital_expenditures",
                "interest_expense",
                "stock_based_compensation",
            },
        }
        dependencies = direct_dependencies.get(field, {field})
        return dependencies.issubset(reconciled_metrics)
    summed_income = {"revenue", "gross_profit", "cost_of_revenue", "operating_income", "net_income", "ebitda", "interest_expense"}
    summed_cash = {
        "cash_from_operations",
        "capital_expenditures",
        "free_cash_flow",
        "fcff",
        "fcff_after_sbc",
        "depreciation_amortization",
        "stock_based_compensation",
        "common_stock_repurchased",
    }
    snapshot_balance = {
        "cash",
        "total_debt",
        "short_term_debt",
        "long_term_debt",
        "total_equity",
        "total_assets",
        "net_receivables",
        "inventory",
        "goodwill_and_intangibles",
        "preferred_stock",
        "minority_interest",
        "unfunded_pension_liability",
        "lease_liabilities_not_in_debt",
        "non_operating_investments",
        "invested_capital",
    }
    points: list[dict[str, Any]] = []
    for field in sorted(summed_income | summed_cash | snapshot_balance | {"diluted_shares"}):
        value = ttm_row.get(field)
        if not _has_value(value):
            continue
        if field in summed_income:
            source_ids = [quarterly_source_ids["income"]]
            formula = "sum of four discrete sequential quarterly income-statement observations"
            if provider_reconciled and metric_is_supported(field, provider_reconciled_metrics):
                source_ids.append(TTM_STATEMENT_SOURCE_IDS["income"])
                formula += "; reconciled to the provider's explicit TTM income statement"
            if metric_is_supported(field, sec_reconciled_metrics):
                source_ids.append(SEC_STATEMENT_SOURCE_IDS["income"])
                formula += "; reconciled to SEC year-to-date filing identities"
        elif field in summed_cash:
            source_ids = [quarterly_source_ids["cash_flow"]]
            formula = "sum of four discrete sequential quarterly cash-flow observations"
            if field in {"fcff", "fcff_after_sbc"}:
                source_ids.append(quarterly_source_ids["income"])
                formula += "; after-tax interest reconciled from the income statement"
            if provider_reconciled and metric_is_supported(field, provider_reconciled_metrics):
                source_ids.append(TTM_STATEMENT_SOURCE_IDS["cash_flow"])
                if field in {"fcff", "fcff_after_sbc"}:
                    source_ids.append(TTM_STATEMENT_SOURCE_IDS["income"])
                formula += "; reconciled to the provider's explicit TTM cash-flow statement"
            if metric_is_supported(field, sec_reconciled_metrics):
                source_ids.append(SEC_STATEMENT_SOURCE_IDS["cash_flow"])
                if field in {"fcff", "fcff_after_sbc"}:
                    source_ids.append(SEC_STATEMENT_SOURCE_IDS["income"])
                formula += "; reconciled to SEC year-to-date filing identities"
        elif field == "diluted_shares":
            source_ids = [quarterly_source_ids["income"]]
            formula = "average diluted shares across four discrete sequential quarters"
            if provider_reconciled and metric_is_supported(field, provider_reconciled_metrics):
                source_ids.append(TTM_STATEMENT_SOURCE_IDS["income"])
                formula += "; reconciled to the provider's explicit TTM income statement"
            if metric_is_supported(field, sec_reconciled_metrics):
                source_ids.append(SEC_STATEMENT_SOURCE_IDS["income"])
                formula += "; reconciled to SEC year-to-date filing identities"
        else:
            if field == "unfunded_pension_liability" and ttm_row.get("unfunded_pension_liability_source_id"):
                source_ids = [str(ttm_row["unfunded_pension_liability_source_id"])]
                formula = str(
                    ttm_row.get("unfunded_pension_liability_basis")
                    or "latest sourced net pension obligation"
                )
            else:
                source_ids = [quarterly_source_ids["balance"]]
                formula = "latest balance-sheet observation within the validated four-quarter window"
            if provider_balance_reconciled and metric_is_supported(field, provider_reconciled_metrics):
                source_ids.append(TTM_STATEMENT_SOURCE_IDS["balance"])
                formula += "; reconciled to the provider's explicit TTM balance sheet"
        point = _data_point(f"financials.ttm.{field}", value, "calculated_metric", formula=formula)
        point["source_ids"] = source_ids
        point["quarter_dates"] = quarter_dates
        points.append(point)
    return points


def _valuation_data_points(valuation: dict[str, Any]) -> list[dict[str, Any]]:
    if not valuation.get("available"):
        return []
    points: list[dict[str, Any]] = []
    for scenario in valuation.get("scenarios", []):
        name = scenario.get("name") or "scenario"
        method = scenario.get("method") or valuation.get("primary_method") or "routed valuation"
        for key in ["intrinsic_value_per_share", "equity_value", "terminal_value", "pv_terminal_value"]:
            value = scenario.get(key)
            if _has_value(value):
                points.append(_data_point(f"valuation.scenario.{name}.{key}", value, "calculated_metric", formula=f"{method} scenario calculation"))
        for key, value in (scenario.get("assumptions") or {}).items():
            if _has_value(value):
                points.append(_data_point(f"valuation.scenario.{name}.assumption.{key}", value, "assumption", formula="scenario assumption"))
    reverse = valuation.get("reverse_dcf") or {}
    price_sources = (valuation.get("price_validation") or {}).get("sources", [])
    reverse_price_source_id = "fmp:quote" if "FMP stable quote" in price_sources else ("fmp:prices" if "FMP latest close" in price_sources else "fmp:profile")
    for key in ["implied_revenue_cagr", "current_price", "value_at_floor", "value_at_ceiling"]:
        value = reverse.get(key)
        if _has_value(value):
            tag = "sourced_fact" if key == "current_price" else "calculated_metric"
            points.append(
                _data_point(
                    f"valuation.reverse_dcf.{key}",
                    value,
                    tag,
                    reverse_price_source_id if key == "current_price" else None,
                    None if key == "current_price" else "binary search for growth where DCF value equals current price",
                )
            )
    for key in ["low", "central", "high"]:
        value = (valuation.get("range") or {}).get(key)
        if _has_value(value):
            formula = "archetype-routed scenarios with explicit method cross-check weights"
            points.append(
                _data_point(
                    f"valuation.range.{key}",
                    value,
                    "calculated_metric",
                    formula=formula,
                )
            )
            points.append(
                _data_point(
                    f"valuation_range_{key}",
                    value,
                    "calculated_metric",
                    formula=formula,
                )
            )
    reliability = valuation.get("reliability") or {}
    if _has_value(reliability.get("score")):
        points.append(
            _data_point(
                "valuation.reliability.score",
                reliability.get("score"),
                "calculated_metric",
                formula="data recency, quote reconciliation, method fit, estimate coverage, disagreement, and terminal-value gates",
            )
        )
    for key, value in (valuation.get("multiples") or {}).items():
        if _has_value(value):
            points.append(_data_point(f"valuation.multiples.{key}", value, "calculated_metric", formula=f"{key} deterministic multiple calculation"))
    return points


def _reconcile_statement_rows(primary_rows: list[dict[str, Any]], sec_rows: list[dict[str, Any]]) -> dict[str, Any]:
    tolerances = {
        "revenue": 0.03,
        "net_income": 0.05,
        "cash_from_operations": 0.05,
        "capital_expenditures": 0.08,
        "cash": 0.05,
        "total_debt": 0.08,
        "total_equity": 0.05,
        "diluted_shares": 0.03,
    }

    def by_year(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        output: dict[str, dict[str, Any]] = {}
        for row in rows:
            date = str(row.get("date") or "")
            if len(date) >= 4 and date[:4].isdigit():
                output[date[:4]] = row
        return output

    primary = by_year(primary_rows)
    sec = by_year(sec_rows)
    overlap = sorted(set(primary) & set(sec))[-3:]
    checks: list[dict[str, Any]] = []
    for year in overlap:
        for metric, tolerance in tolerances.items():
            primary_value = _safe_float(primary[year].get(metric))
            sec_value = _safe_float(sec[year].get(metric))
            if primary_value is None or sec_value is None:
                continue
            denominator = max(abs(primary_value), abs(sec_value), 1.0)
            difference = abs(primary_value - sec_value) / denominator
            checks.append(
                {
                    "period": year,
                    "metric": metric,
                    "primary_value": primary_value,
                    "sec_value": sec_value,
                    "difference": difference,
                    "maximum_difference": tolerance,
                    "passed": difference <= tolerance,
                }
            )

    if len(checks) < 6:
        return {
            "status": "insufficient_overlap",
            "passed": False,
            "overlap_periods": overlap,
            "checks": checks,
            "pass_ratio": None,
        }

    pass_ratio = sum(check["passed"] for check in checks) / len(checks)
    latest_year = overlap[-1]
    critical_metrics = {
        "revenue",
        "cash_from_operations",
        "capital_expenditures",
        "cash",
        "total_debt",
        "total_equity",
        "diluted_shares",
    }
    latest_checks = [check for check in checks if check["period"] == latest_year]
    critical_failures = [
        check
        for check in latest_checks
        if check["metric"] in critical_metrics
        and not check["passed"]
    ]
    observed_critical_metrics = {check["metric"] for check in latest_checks}
    missing_critical_metrics = sorted(critical_metrics - observed_critical_metrics)
    passed = pass_ratio >= 0.80 and not critical_failures and not missing_critical_metrics
    return {
        "status": "reconciled" if passed else "mismatch",
        "passed": passed,
        "overlap_periods": overlap,
        "checks": checks,
        "pass_ratio": pass_ratio,
        "critical_failures": critical_failures,
        "missing_critical_metrics": missing_critical_metrics,
    }


def _apply_statement_reconciliation_gate(
    valuation: dict[str, Any],
    reconciliation: dict[str, Any],
    *,
    required: bool,
) -> dict[str, Any]:
    if not valuation:
        return valuation
    reliability = valuation.setdefault("reliability", {})
    readiness_gates = reliability.setdefault("readiness_gates", {})
    passed = not required or reconciliation.get("passed") is True or reconciliation.get("status") == "primary_sec"
    readiness_gates["sec_statement_reconciliation"] = {
        "passed": passed,
        "required": required,
        "observed": reconciliation.get("status") or "unavailable",
        "pass_ratio": reconciliation.get("pass_ratio"),
    }
    blockers = reliability.setdefault("decision_ready_blockers", [])
    if not passed and "sec_statement_reconciliation" not in blockers:
        blockers.append("sec_statement_reconciliation")
    if not passed:
        limitations = reliability.setdefault("limitations", [])
        message = "Los estados normalizados no están conciliados numéricamente con SEC/XBRL."
        if message not in limitations:
            limitations.append(message)
        reliability["score"] = min(_safe_float(reliability.get("score")) or 0.0, 0.64)
        required_reconciliation_failure = required and not passed
        if required_reconciliation_failure:
            reliability["score"] = min(_safe_float(reliability.get("score")) or 0.0, 0.39)
            reliability["status"] = "blocked"
            reliability["usable"] = False
            valuation["available"] = False
            valuation["status"] = "not_decision_ready"
            valuation["primary_method"] = None
            valuation["cash_flow_basis"] = None
            valuation["range"] = {"low": None, "central": None, "high": None}
            valuation["selected_value"] = None
            valuation["scenarios"] = []
            valuation["methods"] = []
            valuation["reverse_dcf"] = {
                "available": False,
                "status": "blocked_statement_reconciliation",
                "reason": "Los estados requeridos no están conciliados con SEC/XBRL.",
                "weight": 0,
            }
            if "market_requirements" in valuation:
                valuation["market_requirements"] = {
                    "available": False,
                    "status": "blocked_statement_reconciliation",
                    "implied_revenue_cagr": None,
                    "implied_revenue_cagr_bound": None,
                    "normalized_margin": None,
                    "discount_rate": None,
                    "terminal_growth": None,
                    "horizon_years": None,
                    "price_context": None,
                    "reference_price": None,
                    "market_data_as_of": None,
                    "currency": None,
                }
        elif valuation.get("status") == "decision_ready":
            valuation["status"] = "research_grade"
            reliability["status"] = "medium"
            reliability["usable"] = True
    return valuation


def _reconcile_current_share_count(
    ticker: str,
    ttm_row: dict[str, Any] | None,
    shares_float: dict[str, Any] | None,
) -> dict[str, Any]:
    """Compare current basic shares with TTM diluted weighted-average shares.

    The two measures are not interchangeable. The current point-in-time basic
    count is used only to catch material scale, issuance, buyback, split, or
    identity problems in the diluted weighted-average valuation input.
    """
    snapshot = shares_float if isinstance(shares_float, dict) else {}
    current_basic = next(
        (
            value
            for value in (
                _safe_float(snapshot.get("outstandingShares")),
                _safe_float(snapshot.get("outstanding_shares")),
                _safe_float(snapshot.get("sharesOutstanding")),
            )
            if value is not None and value > 0
        ),
        None,
    )
    diluted_ttm = _safe_float((ttm_row or {}).get("diluted_shares"))
    if diluted_ttm is not None and diluted_ttm <= 0:
        diluted_ttm = None
    observed_symbol = clean_ticker(snapshot.get("symbol") or "")
    expected_symbol = clean_ticker(ticker)
    identity_matches = not observed_symbol or observed_symbol == expected_symbol

    difference = None
    if current_basic is not None and diluted_ttm is not None:
        difference = abs(current_basic - diluted_ttm) / diluted_ttm

    snapshot_as_of = snapshot.get("as_of") or snapshot.get("date")
    parsed_as_of = pd.to_datetime(snapshot_as_of, errors="coerce", utc=True)
    now_utc = pd.Timestamp.now(tz="UTC")
    age_days = (now_utc.normalize() - parsed_as_of.normalize()).days if not pd.isna(parsed_as_of) else None
    fresh = age_days is not None and 0 <= age_days <= CURRENT_SHARE_COUNT_MAXIMUM_AGE_DAYS

    required = current_basic is not None and diluted_ttm is not None
    if observed_symbol and not identity_matches:
        status = "identity_mismatch"
        passed: bool | None = False
        required = True
    elif required and not fresh:
        status = "stale" if age_days is not None and age_days > CURRENT_SHARE_COUNT_MAXIMUM_AGE_DAYS else "undated_or_invalid"
        passed = False
    elif not required:
        status = "unavailable"
        passed = None
    elif (
        difference is not None
        and difference >= CURRENT_SHARE_COUNT_MAXIMUM_RELATIVE_DIFFERENCE - 1e-12
    ):
        status = "material_mismatch"
        passed = False
    else:
        status = "reconciled"
        passed = True

    return {
        "status": status,
        "passed": passed,
        "required": required,
        "expected_symbol": expected_symbol,
        "observed_symbol": observed_symbol or None,
        "as_of": snapshot_as_of,
        "age_days": age_days,
        "maximum_age_days": CURRENT_SHARE_COUNT_MAXIMUM_AGE_DAYS,
        "fresh": fresh,
        "current_basic_outstanding_shares": current_basic,
        "ttm_weighted_average_diluted_shares": diluted_ttm,
        "relative_difference": difference,
        "maximum_relative_difference": CURRENT_SHARE_COUNT_MAXIMUM_RELATIVE_DIFFERENCE,
        "valuation_denominator_replaced": False,
        "comparison_role": "control_only_basic_point_in_time_vs_diluted_weighted_average",
    }


def _apply_current_share_count_gate(
    valuation: dict[str, Any],
    reconciliation: dict[str, Any],
) -> dict[str, Any]:
    valuation["share_denominator_reconciliation"] = reconciliation
    reliability = valuation.setdefault("reliability", {})
    readiness_gates = reliability.setdefault("readiness_gates", {})
    required = reconciliation.get("required") is True
    passed = reconciliation.get("passed") is not False
    readiness_gates["current_share_count_reconciliation"] = {
        "passed": passed,
        "required": required,
        "observed": reconciliation.get("status"),
        "relative_difference": reconciliation.get("relative_difference"),
        "maximum_relative_difference": reconciliation.get("maximum_relative_difference"),
        "as_of": reconciliation.get("as_of"),
        "age_days": reconciliation.get("age_days"),
        "maximum_age_days": reconciliation.get("maximum_age_days"),
        "comparison_role": reconciliation.get("comparison_role"),
    }
    if passed:
        return valuation

    blockers = reliability.setdefault("decision_ready_blockers", [])
    if "current_share_count_reconciliation" not in blockers:
        blockers.append("current_share_count_reconciliation")
    limitations = reliability.setdefault("limitations", [])
    message = "Las acciones básicas actuales y el promedio diluido TTM difieren de forma material, están desactualizadas o corresponden a otro instrumento."
    if message not in limitations:
        limitations.append(message)
    reliability["score"] = min(_safe_float(reliability.get("score")) or 0.0, 0.39)
    reliability["status"] = "blocked"
    reliability["usable"] = False
    valuation["available"] = False
    valuation["status"] = "not_decision_ready"
    valuation["primary_method"] = None
    valuation["cash_flow_basis"] = None
    valuation["range"] = {"low": None, "central": None, "high": None}
    valuation["selected_value"] = None
    valuation["scenarios"] = []
    valuation["methods"] = []
    valuation["reverse_dcf"] = {
        "available": False,
        "status": "blocked_share_denominator",
        "reason": "El denominador de acciones no está reconciliado.",
        "weight": 0,
    }
    valuation["market_requirements"] = {
        "available": False,
        "status": "blocked_share_denominator",
        "implied_revenue_cagr": None,
        "implied_revenue_cagr_bound": None,
        "normalized_margin": None,
        "discount_rate": None,
        "terminal_growth": None,
        "horizon_years": None,
        "price_context": None,
        "reference_price": None,
        "market_data_as_of": None,
        "currency": None,
        "assets_added": None,
        "obligations_deducted": None,
    }
    return valuation


def _all_statement_families_use_sec(statement_source_ids: dict[str, str]) -> bool:
    return all(
        statement_source_ids.get(statement_key) == SEC_STATEMENT_SOURCE_IDS[statement_key]
        for statement_key in DEFAULT_STATEMENT_SOURCE_IDS
    )


def _build_evidence_coverage(sources: list[dict[str, Any]], data_points: list[dict[str, Any]]) -> dict[str, Any]:
    source_status = {str(source.get("source_id")): source.get("status") for source in sources if source.get("source_id")}
    statement_source_ids = (
        set(DEFAULT_STATEMENT_SOURCE_IDS.values())
        | set(SEC_STATEMENT_SOURCE_IDS.values())
        | set(YFINANCE_STATEMENT_SOURCE_IDS.values())
        | set(YFINANCE_QUARTERLY_STATEMENT_SOURCE_IDS.values())
    )
    source_provider_by_id = {str(source.get("source_id")): str(source.get("provider")) for source in sources if source.get("source_id")}
    source_backed = [
        point
        for point in data_points
        if point.get("claim_tag") == "sourced_fact" and point.get("source_id") and source_status.get(str(point.get("source_id"))) == "ok"
    ]
    sourced_missing = [
        point.get("metric")
        for point in data_points
        if point.get("claim_tag") == "sourced_fact"
        and (not point.get("source_id") or source_status.get(str(point.get("source_id"))) != "ok")
    ]
    formula_missing = [
        point.get("metric")
        for point in data_points
        if point.get("claim_tag") == "calculated_metric" and not point.get("formula")
    ]
    expected_present = {point.get("metric") for point in data_points if _has_value(point.get("normalized_value"))}
    missing_expected = [metric for metric in EXPECTED_EVIDENCE_METRICS if metric not in expected_present]
    covered_expected = len(EXPECTED_EVIDENCE_METRICS) - len(missing_expected)
    score = round((covered_expected / len(EXPECTED_EVIDENCE_METRICS)) * 100) if EXPECTED_EVIDENCE_METRICS else 0
    sec_metadata_ok = source_status.get("sec:submissions") == "ok"
    statement_sources = [
        source
        for source in sources
        if source.get("source_id") in statement_source_ids
        and source.get("status") == "ok"
        and int(source.get("row_count") or 0) > 0
    ]
    used_statement_source_ids = {
        str(point.get("source_id"))
        for point in data_points
        if str(point.get("source_id") or "") in statement_source_ids
        and _has_value(point.get("normalized_value"))
    }
    used_statement_providers = {
        source_provider_by_id.get(source_id)
        for source_id in used_statement_source_ids
        if source_provider_by_id.get(source_id)
    }
    xbrl_crosscheck_sources = [
        source
        for source in statement_sources
        if source.get("source_id") in set(SEC_STATEMENT_SOURCE_IDS.values())
    ]
    xbrl_statement_facts_available = bool(xbrl_crosscheck_sources)
    reconciliation_source = next((source for source in sources if source.get("source_id") == SEC_RECONCILIATION_SOURCE_ID), {})
    reconciliation_status = reconciliation_source.get("reconciliation_status") or "unavailable"
    if not used_statement_source_ids and not statement_sources:
        statement_authority = "No source-backed normalized statements"
        statement_source_provider = None
    elif used_statement_providers == {"sec-edgar"}:
        statement_authority = "SEC Company Facts/XBRL normalized statements"
        statement_source_provider = "sec-edgar"
    elif "sec-edgar" in used_statement_providers and "fmp" in used_statement_providers:
        statement_authority = "Mixed FMP and SEC Company Facts normalized statements"
        statement_source_provider = "mixed"
    elif "sec-edgar" in used_statement_providers and "yfinance" in used_statement_providers:
        statement_authority = "Yahoo Finance normalized statements with SEC Company Facts support"
        statement_source_provider = "mixed"
    elif "fmp" in used_statement_providers and "yfinance" in used_statement_providers:
        statement_authority = "Mixed FMP and Yahoo Finance normalized statements"
        statement_source_provider = "mixed"
    elif used_statement_providers == {"fmp"} and reconciliation_status == "reconciled":
        statement_authority = "FMP normalized statements reconciled to SEC Company Facts/XBRL"
        statement_source_provider = "fmp"
    elif used_statement_providers == {"fmp"} and xbrl_statement_facts_available:
        statement_authority = "FMP normalized statements; SEC Company Facts available but not reconciled"
        statement_source_provider = "fmp"
    elif used_statement_providers == {"yfinance"} and xbrl_statement_facts_available:
        statement_authority = "Yahoo Finance normalized statements; SEC Company Facts available for review"
        statement_source_provider = "yfinance"
    elif used_statement_providers == {"yfinance"}:
        statement_authority = "Yahoo Finance normalized statements"
        statement_source_provider = "yfinance"
    elif sec_metadata_ok:
        statement_authority = "FMP normalized statements; SEC metadata only"
        statement_source_provider = "fmp"
    else:
        statement_authority = "FMP normalized statements without SEC metadata cross-check"
        statement_source_provider = "fmp"
    status = "pass" if score >= 85 and not sourced_missing and not formula_missing else "partial"
    if score < 60 or sourced_missing:
        status = "needs_attention"
    return {
        "status": status,
        "score": score,
        "expected_metrics": len(EXPECTED_EVIDENCE_METRICS),
        "covered_expected_metrics": covered_expected,
        "missing_expected_metrics": missing_expected,
        "total_data_points": len(data_points),
        "source_backed_points": len(source_backed),
        "calculated_points": len([point for point in data_points if point.get("claim_tag") == "calculated_metric"]),
        "assumption_points": len([point for point in data_points if point.get("claim_tag") == "assumption"]),
        "interpretation_points": len([point for point in data_points if point.get("claim_tag") == "interpretation"]),
        "uncertainty_points": len([point for point in data_points if point.get("claim_tag") == "uncertainty"]),
        "sourced_points_missing_ok_source": sourced_missing,
        "calculated_points_missing_formula": formula_missing,
        "ok_source_records": len([source for source in sources if source.get("status") == "ok"]),
        "error_source_records": len([source for source in sources if source.get("status") == "error"]),
        "unavailable_source_records": len([source for source in sources if source.get("status") == "unavailable"]),
        "sec_metadata_available": sec_metadata_ok,
        "statement_source_provider": statement_source_provider,
        "statement_authority": statement_authority,
        "statement_source_ids": sorted(used_statement_source_ids),
        "statement_crosscheck_source_ids": [str(source.get("source_id")) for source in xbrl_crosscheck_sources if source.get("source_id")],
        "xbrl_statement_facts_available": xbrl_statement_facts_available,
        "statement_reconciliation_status": reconciliation_status,
        "statement_reconciliation_pass_ratio": reconciliation_source.get("pass_ratio"),
    }


def _load_fmp_payloads(ticker: str, paths: PathConfigLike, fmp_client: FMPClient | None) -> tuple[dict[str, Any], dict[str, pd.DataFrame], list[dict[str, Any]]]:
    sources: list[dict[str, Any]] = []
    frames = {
        "income": pd.DataFrame(),
        "cash_flow": pd.DataFrame(),
        "balance": pd.DataFrame(),
        "prices": pd.DataFrame(),
        "income_quarterly": pd.DataFrame(),
        "cash_flow_quarterly": pd.DataFrame(),
        "balance_quarterly": pd.DataFrame(),
        "income_ttm": pd.DataFrame(),
        "cash_flow_ttm": pd.DataFrame(),
        "balance_ttm": pd.DataFrame(),
        "analyst_estimates": pd.DataFrame(),
        "shares_float": {},
        "quote": {},
        "key_metrics_ttm": {},
        "ratios_ttm": {},
    }
    profile: dict[str, Any] = {}

    if fmp_client is None:
        sources.append(_source_record("fmp:client", "fmp", "env:FMP_API_KEY", "unavailable", error="FMP API key is not configured in this runtime."))
        return profile, frames, sources

    loaders = [
        ("profile", "fmp:profile", f"profile/{ticker}", lambda: fmp_client.get_profile(ticker)),
        ("quote", "fmp:quote", f"quote/{ticker}", lambda: fmp_client.get_quote(ticker) if hasattr(fmp_client, "get_quote") else {}),
        (
            "shares_float",
            "fmp:shares-float",
            f"shares-float?symbol={ticker}",
            lambda: fmp_client.get_shares_float(ticker) if hasattr(fmp_client, "get_shares_float") else {},
        ),
        (
            "income",
            "fmp:income:annual",
            f"income-statement/{ticker}?period=annual&limit=10",
            lambda: fmp_client.get_income_statements(ticker, period="annual", limit=10),
        ),
        (
            "cash_flow",
            "fmp:cash-flow:annual",
            f"cash-flow-statement/{ticker}?period=annual&limit=10",
            lambda: fmp_client.get_cash_flow_statements(ticker, period="annual", limit=10),
        ),
        (
            "balance",
            "fmp:balance:annual",
            f"balance-sheet-statement/{ticker}?period=annual&limit=10",
            lambda: fmp_client.get_balance_sheet_statements(ticker, period="annual", limit=10),
        ),
        (
            "prices",
            "fmp:prices",
            f"historical-price-eod/full?symbol={ticker}",
            lambda: fmp_client.get_historical_prices(ticker),
        ),
        (
            "income_quarterly",
            "fmp:income:quarterly",
            f"income-statement/{ticker}?period=quarter&limit=8",
            lambda: fmp_client.get_income_statements(ticker, period="quarter", limit=8),
        ),
        (
            "cash_flow_quarterly",
            "fmp:cash-flow:quarterly",
            f"cash-flow-statement/{ticker}?period=quarter&limit=8",
            lambda: fmp_client.get_cash_flow_statements(ticker, period="quarter", limit=8),
        ),
        (
            "balance_quarterly",
            "fmp:balance:quarterly",
            f"balance-sheet-statement/{ticker}?period=quarter&limit=8",
            lambda: fmp_client.get_balance_sheet_statements(ticker, period="quarter", limit=8),
        ),
        (
            "income_ttm",
            TTM_STATEMENT_SOURCE_IDS["income"],
            f"income-statement-ttm/{ticker}",
            lambda: fmp_client.get_income_statement_ttm(ticker) if hasattr(fmp_client, "get_income_statement_ttm") else pd.DataFrame(),
        ),
        (
            "cash_flow_ttm",
            TTM_STATEMENT_SOURCE_IDS["cash_flow"],
            f"cash-flow-statement-ttm/{ticker}",
            lambda: fmp_client.get_cash_flow_statement_ttm(ticker) if hasattr(fmp_client, "get_cash_flow_statement_ttm") else pd.DataFrame(),
        ),
        (
            "balance_ttm",
            TTM_STATEMENT_SOURCE_IDS["balance"],
            f"balance-sheet-statement-ttm/{ticker}",
            lambda: fmp_client.get_balance_sheet_statement_ttm(ticker) if hasattr(fmp_client, "get_balance_sheet_statement_ttm") else pd.DataFrame(),
        ),
        (
            "analyst_estimates",
            "fmp:analyst-estimates",
            f"analyst-estimates/{ticker}?period=annual&limit=10",
            lambda: fmp_client.get_analyst_estimates(ticker, period="annual", limit=10) if hasattr(fmp_client, "get_analyst_estimates") else pd.DataFrame(),
        ),
        (
            "key_metrics_ttm",
            "fmp:key-metrics-ttm",
            f"key-metrics-ttm/{ticker}",
            lambda: fmp_client.get_key_metrics_ttm(ticker) if hasattr(fmp_client, "get_key_metrics_ttm") else {},
        ),
        (
            "ratios_ttm",
            "fmp:ratios-ttm",
            f"ratios-ttm/{ticker}",
            lambda: fmp_client.get_ratios_ttm(ticker) if hasattr(fmp_client, "get_ratios_ttm") else {},
        ),
    ]

    for key, source_id, endpoint, loader in loaders:
        try:
            payload = loader()
            if key == "profile":
                profile = payload if isinstance(payload, dict) else {}
                status = "ok" if profile else "unavailable"
                sources.append(_source_record(source_id, "fmp", endpoint, status, row_count=1 if profile else 0, as_of=profile.get("as_of")))
            elif key in {"quote", "shares_float", "key_metrics_ttm", "ratios_ttm"}:
                record = payload if isinstance(payload, dict) else {}
                frames[key] = record
                status = "ok" if record else "unavailable"
                sources.append(_source_record(source_id, "fmp", endpoint, status, row_count=1 if record else 0, as_of=record.get("as_of")))
            else:
                frame = payload if isinstance(payload, pd.DataFrame) else pd.DataFrame(payload)
                provider_snapshot_at = None
                forecast_through = None
                if key == "analyst_estimates" and not frame.empty:
                    frame = frame.copy()
                    provider_snapshot_at = _now_iso()
                    frame["providerSnapshotAt"] = provider_snapshot_at
                    frame["sourceFamily"] = "FMP"
                    frame["provenanceBasis"] = "current_provider_snapshot_retrieved_at"
                    profile_currency = str(profile.get("currency") or "").upper().strip()
                    if profile_currency:
                        if "currency" not in frame.columns:
                            frame["currency"] = profile_currency
                        else:
                            explicit_currency = frame["currency"].notna() & frame["currency"].astype(str).str.strip().ne("")
                            frame["currency"] = frame["currency"].where(explicit_currency, profile_currency)
                        frame["currencyBasis"] = "fmp_profile_listing_currency_cross_checked_against_statements"
                frames[key] = frame
                status = "ok" if not frame.empty else "unavailable"
                as_of = None
                if not frame.empty and "date" in frame.columns:
                    parsed_dates = pd.to_datetime(frame["date"], errors="coerce").dropna()
                    as_of = str(parsed_dates.max().date()) if not parsed_dates.empty else None
                    forecast_through = as_of if key == "analyst_estimates" else None
                sources.append(
                    _source_record(
                        source_id,
                        "fmp",
                        endpoint,
                        status,
                        row_count=int(len(frame)),
                        as_of=provider_snapshot_at or as_of,
                        forecast_through=forecast_through,
                    )
                )
        except Exception as exc:  # noqa: BLE001
            sources.append(_source_record(source_id, "fmp", endpoint, "error", error=str(exc)))

    return profile, frames, sources


def _yfinance_value(payload: Any, *keys: str) -> Any:
    for key in keys:
        try:
            value = payload.get(key) if hasattr(payload, "get") else getattr(payload, key)
        except (AttributeError, KeyError, TypeError, ValueError):
            value = None
        if _has_value(value):
            return value
    return None


def _normalized_yfinance_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _yfinance_statement_frame(
    payload: Any,
    aliases: dict[str, tuple[str, ...]],
    *,
    period: str,
    currency: str | None,
) -> pd.DataFrame:
    frame = payload.copy() if isinstance(payload, pd.DataFrame) else pd.DataFrame(payload)
    if frame.empty:
        return pd.DataFrame()
    labels = {_normalized_yfinance_label(label): label for label in frame.index}
    rows: list[dict[str, Any]] = []
    for column in frame.columns:
        parsed_date = pd.to_datetime(column, errors="coerce")
        if pd.isna(parsed_date):
            continue
        row: dict[str, Any] = {
            "date": str(parsed_date.date()),
            "calendarYear": int(parsed_date.year),
            "fiscalYear": int(parsed_date.year),
            "period": period,
        }
        if currency:
            row["reportedCurrency"] = currency
        for target, candidates in aliases.items():
            source_label = next(
                (labels[_normalized_yfinance_label(candidate)] for candidate in candidates if _normalized_yfinance_label(candidate) in labels),
                None,
            )
            if source_label is None:
                continue
            value = _safe_float(frame.at[source_label, column])
            if value is not None:
                row[target] = value
        rows.append(row)
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True) if rows else pd.DataFrame()


def _yfinance_history_frame(payload: Any, ticker: str, currency: str | None) -> pd.DataFrame:
    frame = payload.copy() if isinstance(payload, pd.DataFrame) else pd.DataFrame(payload)
    if frame.empty:
        return pd.DataFrame()
    close_column = next(
        (column for column in frame.columns if _normalized_yfinance_label(column) in {"close", "adjclose"}),
        None,
    )
    if close_column is None:
        return pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for index, value in frame[close_column].items():
        parsed_date = pd.to_datetime(index, errors="coerce")
        close = _safe_float(value)
        if pd.isna(parsed_date) or close is None or close <= 0:
            continue
        rows.append(
            {
                "date": str(parsed_date.date()),
                "close": close,
                "symbol": ticker,
                "currency": currency,
                "source_family": "yfinance",
            }
        )
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True) if rows else pd.DataFrame()


def _load_yfinance_payloads(
    ticker: str,
    yfinance_factory: Any | None,
    *,
    include_statements: bool,
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    """Load an auditable Yahoo Finance fallback without weakening valuation gates."""

    sources: list[dict[str, Any]] = []
    frames: dict[str, Any] = {
        "income": pd.DataFrame(),
        "cash_flow": pd.DataFrame(),
        "balance": pd.DataFrame(),
        "prices": pd.DataFrame(),
        "income_quarterly": pd.DataFrame(),
        "cash_flow_quarterly": pd.DataFrame(),
        "balance_quarterly": pd.DataFrame(),
        "shares_float": {},
        "quote": {},
    }
    factory = yfinance_factory
    if factory is None:
        try:
            import yfinance as yf  # type: ignore

            factory = yf.Ticker
        except Exception as exc:  # noqa: BLE001
            sources.append(
                _source_record(
                    "yfinance:client",
                    "yfinance",
                    "python:yfinance",
                    "unavailable",
                    error=f"Yahoo Finance fallback is unavailable: {exc}",
                )
            )
            return {}, frames, sources

    try:
        instrument = factory(ticker)
    except Exception as exc:  # noqa: BLE001
        sources.append(_source_record("yfinance:client", "yfinance", f"Ticker/{ticker}", "error", error=str(exc)))
        return {}, frames, sources

    info: Any = {}
    fast_info: Any = {}
    info_error: Exception | None = None
    fast_info_error: Exception | None = None
    try:
        info = instrument.info or {}
    except Exception as exc:  # noqa: BLE001
        info_error = exc
    try:
        fast_info = instrument.fast_info or {}
    except Exception as exc:  # noqa: BLE001
        fast_info_error = exc

    currency = str(
        _yfinance_value(fast_info, "currency")
        or _yfinance_value(info, "currency", "financialCurrency")
        or ""
    ).upper().strip() or None
    history = pd.DataFrame()
    history_error: Exception | None = None
    try:
        history = _yfinance_history_frame(
            instrument.history(period="1mo", interval="1d", auto_adjust=False, actions=False),
            ticker,
            currency,
        )
    except Exception as exc:  # noqa: BLE001
        history_error = exc
    frames["prices"] = history
    history_as_of = None if history.empty else str(history["date"].max())

    price = _safe_float(
        _yfinance_value(fast_info, "last_price", "lastPrice")
        or _yfinance_value(info, "currentPrice", "regularMarketPrice")
    )
    market_cap = _safe_float(
        _yfinance_value(fast_info, "market_cap", "marketCap")
        or _yfinance_value(info, "marketCap")
    )
    shares = _safe_float(
        _yfinance_value(fast_info, "shares", "sharesOutstanding")
        or _yfinance_value(info, "sharesOutstanding")
    )
    exchange = str(
        _yfinance_value(fast_info, "exchange")
        or _yfinance_value(info, "exchange", "fullExchangeName")
        or ""
    ).upper().strip() or None
    profile = {
        "symbol": str(_yfinance_value(info, "symbol") or ticker).upper(),
        "companyName": _yfinance_value(info, "longName", "shortName") or ticker,
        "sector": _yfinance_value(info, "sector"),
        "industry": _yfinance_value(info, "industry"),
        "country": _yfinance_value(info, "country"),
        "currency": currency,
        "exchangeShortName": exchange,
        "price": price,
        "marketCap": market_cap,
        "mktCap": market_cap,
        "beta": _safe_float(_yfinance_value(info, "beta")),
        "description": _yfinance_value(info, "longBusinessSummary"),
        "source_family": "yfinance",
    }
    profile = {key: value for key, value in profile.items() if _has_value(value)}
    profile_has_evidence = any(
        _has_value(_yfinance_value(info, key))
        for key in ("longName", "shortName", "sector", "industry", "country", "currency", "marketCap")
    )
    sources.append(
        _source_record(
            "yfinance:profile",
            "yfinance",
            f"Ticker/{ticker}/info",
            "ok" if profile_has_evidence else ("error" if info_error else "unavailable"),
            row_count=1 if profile_has_evidence else 0,
            error=str(info_error) if info_error else None,
        )
    )

    quote = {
        "symbol": ticker,
        "price": price,
        "marketCap": market_cap,
        "currency": currency,
        "exchange": exchange,
        "as_of": history_as_of,
        "source_family": "yfinance",
    }
    quote = {key: value for key, value in quote.items() if _has_value(value)}
    frames["quote"] = quote
    sources.append(
        _source_record(
            "yfinance:quote",
            "yfinance",
            f"Ticker/{ticker}/fast_info",
            "ok" if price is not None else ("error" if fast_info_error else "unavailable"),
            row_count=1 if price is not None else 0,
            as_of=history_as_of,
            error=str(fast_info_error) if fast_info_error else None,
        )
    )
    sources.append(
        _source_record(
            "yfinance:prices",
            "yfinance",
            f"Ticker/{ticker}/history?period=1mo&interval=1d",
            "ok" if not history.empty else ("error" if history_error else "unavailable"),
            row_count=int(len(history)),
            as_of=history_as_of,
            error=str(history_error) if history_error else None,
        )
    )

    shares_float = {
        "symbol": ticker,
        "outstandingShares": shares,
        "floatShares": _safe_float(_yfinance_value(info, "floatShares")),
        "as_of": history_as_of,
        "source_family": "yfinance",
    }
    shares_float = {key: value for key, value in shares_float.items() if _has_value(value)}
    frames["shares_float"] = shares_float
    sources.append(
        _source_record(
            "yfinance:shares-float",
            "yfinance",
            f"Ticker/{ticker}/info#shares",
            "ok" if shares is not None else "unavailable",
            row_count=1 if shares is not None else 0,
            as_of=history_as_of,
        )
    )

    if not include_statements:
        return profile, frames, sources

    income_aliases = {
        "revenue": ("Total Revenue", "Operating Revenue"),
        "grossProfit": ("Gross Profit",),
        "costOfRevenue": ("Cost Of Revenue",),
        "operatingIncome": ("Operating Income",),
        "incomeBeforeTax": ("Pretax Income",),
        "incomeTaxExpense": ("Tax Provision",),
        "netIncome": ("Net Income", "Net Income Common Stockholders"),
        "ebitda": ("EBITDA", "Normalized EBITDA"),
        "interestExpense": ("Interest Expense", "Interest Expense Non Operating"),
        "weightedAverageShsOutDil": ("Diluted Average Shares", "Basic Average Shares"),
    }
    cash_flow_aliases = {
        "netCashProvidedByOperatingActivities": ("Operating Cash Flow", "Cash Flow From Continuing Operating Activities"),
        "capitalExpenditure": ("Capital Expenditure",),
        "depreciationAndAmortization": ("Depreciation And Amortization", "Depreciation Amortization Depletion"),
        "stockBasedCompensation": ("Stock Based Compensation",),
        "commonStockRepurchased": ("Repurchase Of Capital Stock",),
    }
    balance_aliases = {
        "cashAndCashEquivalents": ("Cash And Cash Equivalents", "Cash Cash Equivalents"),
        "cashAndShortTermInvestments": ("Cash Cash Equivalents And Short Term Investments",),
        "shortTermInvestments": ("Other Short Term Investments", "Financial Assets"),
        "totalDebt": ("Total Debt",),
        "shortTermDebt": ("Current Debt", "Current Debt And Capital Lease Obligation"),
        "longTermDebt": ("Long Term Debt", "Long Term Debt And Capital Lease Obligation"),
        "totalStockholdersEquity": ("Stockholders Equity", "Common Stock Equity"),
        "totalAssets": ("Total Assets",),
        "netReceivables": ("Receivables", "Accounts Receivable"),
        "inventory": ("Inventory",),
        "goodwill": ("Goodwill",),
        "intangibleAssets": ("Other Intangible Assets Net",),
        "preferredStock": ("Preferred Stock Equity",),
        "minorityInterest": ("Minority Interest",),
    }
    loaders = [
        ("income", "yfinance:income:annual", "income_stmt", "FY", income_aliases),
        ("cash_flow", "yfinance:cash-flow:annual", "cash_flow", "FY", cash_flow_aliases),
        ("balance", "yfinance:balance:annual", "balance_sheet", "FY", balance_aliases),
        ("income_quarterly", "yfinance:income:quarterly", "quarterly_income_stmt", "Q", income_aliases),
        ("cash_flow_quarterly", "yfinance:cash-flow:quarterly", "quarterly_cash_flow", "Q", cash_flow_aliases),
        ("balance_quarterly", "yfinance:balance:quarterly", "quarterly_balance_sheet", "Q", balance_aliases),
    ]
    for key, source_id, attribute, period, aliases in loaders:
        endpoint = f"Ticker/{ticker}/{attribute}?period={'annual' if period == 'FY' else 'quarterly'}"
        try:
            frame = _yfinance_statement_frame(
                getattr(instrument, attribute),
                aliases,
                period=period,
                currency=currency,
            )
            frames[key] = frame
            as_of = None if frame.empty else str(frame["date"].max())
            sources.append(
                _source_record(
                    source_id,
                    "yfinance",
                    endpoint,
                    "ok" if not frame.empty else "unavailable",
                    row_count=int(len(frame)),
                    as_of=as_of,
                )
            )
        except Exception as exc:  # noqa: BLE001
            sources.append(_source_record(source_id, "yfinance", endpoint, "error", row_count=0, error=str(exc)))

    return profile, frames, sources


def _load_sec_filings(ticker: str, sec_client: SECEdgarClient | None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    source_id = "sec:submissions"
    endpoint = f"submissions/CIK{{resolved_from_{ticker}}}.json"
    if sec_client is None:
        return [], [
            _source_record(
                source_id,
                "sec-edgar",
                "env:SEC_USER_AGENT",
                "unavailable",
                error="SEC_USER_AGENT is not configured; EDGAR metadata was not requested.",
            )
        ]
    try:
        filings = sec_client.get_recent_filings(ticker)
        return filings, [_source_record(source_id, "sec-edgar", endpoint, "ok", row_count=len(filings))]
    except Exception as exc:  # noqa: BLE001
        return [], [_source_record(source_id, "sec-edgar", endpoint, "error", error=str(exc))]


def _sec_unit_priority(target: str) -> tuple[str, ...]:
    return ("shares",) if target == "weightedAverageShsOutDil" else ("USD",)


def _sec_us_gaap_facts(company_facts: dict[str, Any]) -> dict[str, Any]:
    facts = company_facts.get("facts") if isinstance(company_facts, dict) else {}
    us_gaap = facts.get("us-gaap") if isinstance(facts, dict) else {}
    return us_gaap if isinstance(us_gaap, dict) else {}


def _sec_concept_values_by_date(
    company_facts: dict[str, Any],
    concept_names: list[str],
    *,
    unit_priority: tuple[str, ...],
    instantaneous: bool = False,
) -> dict[str, dict[str, Any]]:
    us_gaap = _sec_us_gaap_facts(company_facts)
    values: dict[str, dict[str, Any]] = {}
    for concept in concept_names:
        concept_payload = us_gaap.get(concept) or {}
        units = concept_payload.get("units") if isinstance(concept_payload, dict) else {}
        if not isinstance(units, dict):
            continue
        unit = next((candidate for candidate in unit_priority if candidate in units), None)
        if unit is None:
            continue
        for fact in units.get(unit) or []:
            if not isinstance(fact, dict):
                continue
            form = str(fact.get("form") or "").upper()
            accepted_forms = SEC_ANNUAL_FORMS | ({"10-Q", "10-Q/A"} if instantaneous else set())
            if form not in accepted_forms:
                continue
            fiscal_period = str(fact.get("fp") or "").upper()
            if not instantaneous and fiscal_period and fiscal_period != "FY":
                continue
            end_date = str(fact.get("end") or "").strip()
            value = _safe_float(fact.get("val"))
            if not end_date or value is None:
                continue
            filed = str(fact.get("filed") or "")
            existing = values.get(end_date)
            if existing is None or (existing.get("concept") == concept and filed >= str(existing.get("filed") or "")):
                values[end_date] = {
                    "value": value,
                    "concept": concept,
                    "unit": unit,
                    "filed": filed,
                    "form": form,
                    "fy": fact.get("fy"),
                }
    return values


def _sec_company_facts_frames(company_facts: dict[str, Any]) -> tuple[dict[str, pd.DataFrame], dict[str, list[str]]]:
    frames: dict[str, pd.DataFrame] = {}
    targets_covered: dict[str, list[str]] = {}
    for statement_key, concept_map in SEC_CONCEPTS.items():
        rows_by_date: dict[str, dict[str, Any]] = {}
        covered: list[str] = []
        for target, concepts in concept_map.items():
            values_by_date = _sec_concept_values_by_date(
                company_facts,
                concepts,
                unit_priority=_sec_unit_priority(target),
                instantaneous=False,
            )
            if values_by_date:
                covered.append(target)
            for end_date, fact in values_by_date.items():
                rows_by_date.setdefault(end_date, {"date": end_date})
                rows_by_date[end_date][target] = fact["value"]
                unit = str(fact.get("unit") or "").upper()
                if statement_key == "income" and len(unit) == 3 and unit.isalpha() and unit != "SHARES":
                    rows_by_date[end_date]["reportedCurrency"] = unit
        if statement_key == "balance":
            for row in rows_by_date.values():
                if row.get("goodwillAndIntangibleAssets") is None:
                    goodwill = _safe_float(row.get("goodwill"))
                    intangible_assets = _safe_float(row.get("intangibleAssets"))
                    if goodwill is not None or intangible_assets is not None:
                        row["goodwillAndIntangibleAssets"] = (goodwill or 0.0) + (intangible_assets or 0.0)
                pension_obligation = _safe_float(row.get("pensionBenefitObligation"))
                pension_assets = _safe_float(row.get("pensionPlanAssets"))
                if pension_obligation is not None and pension_assets is not None:
                    row["unfundedPensionLiability"] = max(0.0, pension_obligation - pension_assets)
                    row["unfundedPensionLiabilityBasis"] = "benefit_obligation_less_plan_assets"
                    covered.append("unfundedPensionLiability")
                row.pop("goodwill", None)
                row.pop("intangibleAssets", None)
        frame = pd.DataFrame([rows_by_date[date] for date in sorted(rows_by_date)])
        frames[statement_key] = frame
        targets_covered[statement_key] = sorted(set(covered))

    pension_rows_by_date: dict[str, dict[str, Any]] = {}
    for target in ("pensionBenefitObligation", "pensionPlanAssets"):
        values_by_date = _sec_concept_values_by_date(
            company_facts,
            SEC_CONCEPTS["balance"][target],
            unit_priority=_sec_unit_priority(target),
            instantaneous=True,
        )
        for end_date, fact in values_by_date.items():
            pension_rows_by_date.setdefault(end_date, {"date": end_date})
            pension_rows_by_date[end_date][target] = fact["value"]
    for row in pension_rows_by_date.values():
        pension_obligation = _safe_float(row.get("pensionBenefitObligation"))
        pension_assets = _safe_float(row.get("pensionPlanAssets"))
        if pension_obligation is not None and pension_assets is not None:
            row["unfundedPensionLiability"] = max(0.0, pension_obligation - pension_assets)
            row["unfundedPensionLiabilityBasis"] = "benefit_obligation_less_plan_assets"
    pension_rows = [
        pension_rows_by_date[date]
        for date in sorted(pension_rows_by_date)
        if _safe_float(pension_rows_by_date[date].get("unfundedPensionLiability")) is not None
    ]
    frames["balance_enrichment"] = pd.DataFrame(pension_rows)
    if pension_rows:
        targets_covered["balance"] = sorted(
            set([*targets_covered.get("balance", []), "unfundedPensionLiability"])
        )
    return frames, targets_covered


def _load_sec_company_facts(
    ticker: str,
    sec_client: SECEdgarClient | None,
) -> tuple[dict[str, pd.DataFrame], list[dict[str, Any]], dict[str, Any]]:
    frames = {
        "income": pd.DataFrame(),
        "cash_flow": pd.DataFrame(),
        "balance": pd.DataFrame(),
    }
    endpoint = f"api/xbrl/companyfacts/CIK{{resolved_from_{ticker}}}.json"
    if sec_client is None:
        return frames, [
            _source_record(
                source_id,
                "sec-edgar",
                "env:SEC_USER_AGENT",
                "unavailable",
                error="SEC_USER_AGENT is not configured; XBRL company facts were not requested.",
            )
            for source_id in SEC_STATEMENT_SOURCE_IDS.values()
        ], {}
    try:
        company_facts = sec_client.get_company_facts(ticker)
        if not company_facts:
            return frames, [
                _source_record(source_id, "sec-edgar", endpoint, "unavailable", row_count=0, error="No SEC company facts payload was returned.")
                for source_id in SEC_STATEMENT_SOURCE_IDS.values()
            ], {}
        frames, targets_covered = _sec_company_facts_frames(company_facts)
        sources: list[dict[str, Any]] = []
        for statement_key, source_id in SEC_STATEMENT_SOURCE_IDS.items():
            frame = frames.get(statement_key, pd.DataFrame())
            status = "ok" if not frame.empty else "unavailable"
            sources.append(
                _source_record(
                    source_id,
                    "sec-edgar",
                    endpoint,
                    status,
                    row_count=int(len(frame)),
                    targets_covered=targets_covered.get(statement_key, []),
                    error=None if status == "ok" else "No annual SEC facts mapped to this normalized statement.",
                )
            )
        return frames, sources, company_facts
    except Exception as exc:  # noqa: BLE001
        return frames, [
            _source_record(source_id, "sec-edgar", endpoint, "error", error=str(exc))
            for source_id in SEC_STATEMENT_SOURCE_IDS.values()
        ], {}


def _enrich_balance_frames_with_sec(
    frames: dict[str, Any],
    sec_frames: dict[str, pd.DataFrame],
) -> list[dict[str, Any]]:
    sec_balance = sec_frames.get("balance_enrichment", pd.DataFrame())
    if not isinstance(sec_balance, pd.DataFrame) or sec_balance.empty:
        sec_balance = sec_frames.get("balance", pd.DataFrame())
    if not isinstance(sec_balance, pd.DataFrame) or sec_balance.empty:
        return []
    if "date" not in sec_balance.columns or "unfundedPensionLiability" not in sec_balance.columns:
        return []

    sec_rows = sec_balance.copy()
    sec_rows["_parsed_date"] = pd.to_datetime(sec_rows["date"], errors="coerce")
    sec_rows = sec_rows.dropna(subset=["_parsed_date"]).sort_values("_parsed_date")
    enrichments: list[dict[str, Any]] = []
    for frame_key in ("balance", "balance_quarterly", "balance_ttm"):
        frame = frames.get(frame_key)
        if not isinstance(frame, pd.DataFrame) or frame.empty or "date" not in frame.columns:
            continue
        enriched = frame.copy()
        for column in (
            "unfundedPensionLiability",
            "unfundedPensionLiabilityBasis",
            "unfundedPensionLiabilityAsOf",
            "unfundedPensionLiabilitySourceId",
        ):
            if column not in enriched.columns:
                enriched[column] = None
        for index, raw in enriched.iterrows():
            if _safe_float(raw.get("unfundedPensionLiability")) is not None:
                continue
            target_date = pd.to_datetime(raw.get("date"), errors="coerce")
            if pd.isna(target_date):
                continue
            candidates = sec_rows[
                (sec_rows["_parsed_date"] <= target_date)
                & ((target_date - sec_rows["_parsed_date"]).dt.days <= 550)
            ]
            candidates = candidates[
                candidates["unfundedPensionLiability"].map(lambda value: _safe_float(value) is not None)
            ]
            if candidates.empty:
                continue
            source = candidates.iloc[-1]
            as_of = str(source["_parsed_date"].date())
            basis = str(source.get("unfundedPensionLiabilityBasis") or "sec_explicit_net_liability")
            value = float(_safe_float(source.get("unfundedPensionLiability")) or 0.0)
            enriched.at[index, "unfundedPensionLiability"] = value
            enriched.at[index, "unfundedPensionLiabilityBasis"] = basis
            enriched.at[index, "unfundedPensionLiabilityAsOf"] = as_of
            enriched.at[index, "unfundedPensionLiabilitySourceId"] = SEC_STATEMENT_SOURCE_IDS["balance"]
            enrichments.append(
                {
                    "frame": frame_key,
                    "date": str(target_date.date()),
                    "field": "unfunded_pension_liability",
                    "value": value,
                    "source_as_of": as_of,
                    "basis": basis,
                    "carry_forward_days": int((target_date - source["_parsed_date"]).days),
                }
            )
        frames[frame_key] = enriched
    return enrichments


class PathConfigLike:
    cache_root: Path


def _normalize_financials(frames: dict[str, pd.DataFrame], tax_rate: float) -> list[dict[str, Any]]:
    income = frames["income"].copy()
    cash_flow = frames["cash_flow"].copy()
    balance = frames["balance"].copy()

    def normalized_year(value: Any) -> tuple[int | None, bool]:
        if value in (None, ""):
            return None, False
        try:
            if pd.isna(value):
                return None, False
        except (TypeError, ValueError):
            pass
        try:
            numeric = float(str(value).strip())
        except (TypeError, ValueError):
            return None, True
        if not math.isfinite(numeric) or not numeric.is_integer():
            return None, True
        year = int(numeric)
        return (year, False) if 1900 <= year <= 2200 else (None, True)

    def project(frame: pd.DataFrame, fields: dict[str, list[str]], *, family: str) -> pd.DataFrame:
        metadata_columns = [
            f"{family}_fiscal_year",
            f"{family}_calendar_year",
            f"{family}_period",
            f"{family}_fiscal_year_invalid",
            f"{family}_calendar_year_invalid",
            f"{family}_statement_duplicate_mismatch",
            f"{family}_statement_duplicate_conflicting_fields",
            f"{family}_statement_duplicate_count",
        ]
        if frame.empty or "date" not in frame.columns:
            return pd.DataFrame(columns=["date", *metadata_columns, *fields.keys()])
        rows = []
        for _, raw in frame.iterrows():
            row = {"date": str(pd.to_datetime(raw.get("date")).date())}
            fiscal_year, fiscal_year_invalid = normalized_year(
                _first_existing(raw, ["fiscalYear", "fiscal_year", "fy"])
            )
            calendar_year, calendar_year_invalid = normalized_year(
                _first_existing(raw, ["calendarYear", "calendar_year"])
            )
            period_value = _first_existing(raw, ["period", "fiscalPeriod", "fiscal_period", "fp"])
            row.update(
                {
                    f"{family}_fiscal_year": fiscal_year,
                    f"{family}_calendar_year": calendar_year,
                    f"{family}_period": (
                        str(period_value).upper().strip()
                        if period_value not in (None, "")
                        else None
                    ),
                    f"{family}_fiscal_year_invalid": fiscal_year_invalid,
                    f"{family}_calendar_year_invalid": calendar_year_invalid,
                }
            )
            for target, names in fields.items():
                value = _first_existing(raw, names)
                if target.endswith("_currency") and value not in (None, ""):
                    row[target] = str(value).upper().strip()
                elif target.endswith(("_basis", "_as_of", "_source_id")) and value not in (None, ""):
                    row[target] = str(value).strip()
                else:
                    row[target] = _safe_float(value)
            rows.append(row)
        projected = pd.DataFrame(rows)
        collapsed: list[dict[str, Any]] = []
        metadata_field_labels = {
            f"{family}_fiscal_year": "fiscal_year",
            f"{family}_calendar_year": "calendar_year",
            f"{family}_period": "period",
            f"{family}_fiscal_year_invalid": "fiscal_year_invalid",
            f"{family}_calendar_year_invalid": "calendar_year_invalid",
        }
        for date, same_date in projected.groupby("date", sort=False):
            collapsed_row: dict[str, Any] = {"date": date}
            conflicts: list[str] = []
            for column in projected.columns:
                if column == "date":
                    continue
                observed: dict[tuple[str, Any], Any] = {}
                for value in same_date[column].tolist():
                    try:
                        if pd.isna(value):
                            continue
                    except (TypeError, ValueError):
                        pass
                    if column.endswith("_invalid"):
                        key = ("boolean", bool(value))
                    elif pd.api.types.is_number(value):
                        key = ("number", float(value))
                    else:
                        key = ("text", str(value))
                    observed.setdefault(key, value)
                if column.endswith("_invalid"):
                    collapsed_row[column] = any(bool(value) for value in observed.values())
                elif len(observed) <= 1:
                    collapsed_row[column] = next(iter(observed.values()), None)
                else:
                    collapsed_row[column] = None
                if len(observed) > 1:
                    conflicts.append(metadata_field_labels.get(column, column))
            collapsed_row[f"{family}_statement_duplicate_mismatch"] = bool(conflicts)
            collapsed_row[f"{family}_statement_duplicate_conflicting_fields"] = sorted(conflicts)
            collapsed_row[f"{family}_statement_duplicate_count"] = int(len(same_date))
            collapsed.append(collapsed_row)
        return pd.DataFrame(collapsed)

    income_n = project(
        income,
        {
            "revenue": ["revenue"],
            "gross_profit": ["grossProfit", "gross_profit"],
            "cost_of_revenue": ["costOfRevenue", "cost_of_revenue"],
            "operating_income": ["operatingIncome", "operating_income"],
            "pretax_income": ["incomeBeforeTax", "income_before_tax"],
            "tax_expense": ["incomeTaxExpense", "income_tax_expense"],
            "net_income": ["netIncome", "net_income"],
            "ebitda": ["ebitda", "EBITDA"],
            "interest_expense": ["interestExpense", "interestExpenseNonOperating", "interest_expense"],
            "diluted_shares": ["weightedAverageShsOutDil", "weighted_average_shares_diluted"],
            "income_currency": ["reportedCurrency", "reported_currency"],
        },
        family="income",
    )
    cash_n = project(
        cash_flow,
        {
            "cash_from_operations": ["netCashProvidedByOperatingActivities", "operatingCashFlow", "cash_from_operations"],
            "capital_expenditures_raw": ["capitalExpenditure", "capital_expenditures"],
            "depreciation_amortization": ["depreciationAndAmortization", "depreciation_amortization"],
            "stock_based_compensation": ["stockBasedCompensation", "stock_based_compensation"],
            "common_stock_repurchased": ["commonStockRepurchased", "common_stock_repurchased"],
            "cash_flow_currency": ["reportedCurrency", "reported_currency"],
        },
        family="cash_flow",
    )
    balance_n = project(
        balance,
        {
            "cash": ["cashAndCashEquivalents", "cash"],
            "cash_and_short_term_investments": ["cashAndShortTermInvestments"],
            "short_term_investments": ["shortTermInvestments"],
            "total_debt": ["totalDebt"],
            "short_term_debt": ["shortTermDebt", "short_term_debt"],
            "long_term_debt": ["longTermDebt", "long_term_debt"],
            "total_equity": ["totalStockholdersEquity", "total_equity"],
            "total_assets": ["totalAssets", "total_assets"],
            "net_receivables": ["netReceivables", "net_receivables"],
            "inventory": ["inventory"],
            "goodwill_and_intangibles": ["goodwillAndIntangibleAssets", "goodwill_and_intangibles"],
            "preferred_stock": ["preferredStock", "preferredStockEquity", "preferred_stock"],
            "minority_interest": ["minorityInterest", "noncontrollingInterestInConsolidatedEntity", "minority_interest"],
            "unfunded_pension_liability": ["unfundedPensionLiability", "unfunded_pension_liability"],
            "unfunded_pension_liability_basis": ["unfundedPensionLiabilityBasis"],
            "unfunded_pension_liability_as_of": ["unfundedPensionLiabilityAsOf"],
            "unfunded_pension_liability_source_id": ["unfundedPensionLiabilitySourceId"],
            "lease_liabilities_not_in_debt": ["leaseLiabilitiesNotInDebt", "lease_liabilities_not_in_debt"],
            "capital_lease_obligations": ["capitalLeaseObligations", "financeLeaseLiabilities", "capital_lease_obligations"],
            "balance_currency": ["reportedCurrency", "reported_currency"],
        },
        family="balance",
    )

    merged = income_n.merge(cash_n, on="date", how="outer").merge(balance_n, on="date", how="outer")
    if merged.empty:
        return []
    merged = merged.sort_values("date")

    rows: list[dict[str, Any]] = []
    for _, row in merged.iterrows():
        def reconciled_years(suffix: str) -> list[int]:
            values: set[int] = set()
            for family in ("income", "cash_flow", "balance"):
                value, invalid = normalized_year(row.get(f"{family}_{suffix}"))
                if value is not None and not invalid:
                    values.add(value)
            return sorted(values)

        def reconciled_texts(suffix: str) -> list[str]:
            values: set[str] = set()
            for family in ("income", "cash_flow", "balance"):
                value = row.get(f"{family}_{suffix}")
                try:
                    if pd.isna(value):
                        continue
                except (TypeError, ValueError):
                    pass
                text = str(value or "").upper().strip()
                if text:
                    values.add(text)
            return sorted(values)

        def has_invalid_metadata() -> bool:
            for family in ("income", "cash_flow", "balance"):
                for suffix in ("fiscal_year_invalid", "calendar_year_invalid"):
                    value = row.get(f"{family}_{suffix}")
                    try:
                        if pd.isna(value):
                            continue
                    except (TypeError, ValueError):
                        pass
                    if bool(value):
                        return True
            return False

        fiscal_years = reconciled_years("fiscal_year")
        calendar_years = reconciled_years("calendar_year")
        periods = reconciled_texts("period")
        fiscal_metadata_mismatch = bool(
            len(fiscal_years) > 1
            or len(calendar_years) > 1
            or len(periods) > 1
            or has_invalid_metadata()
        )
        fiscal_year = (
            fiscal_years[0]
            if len(fiscal_years) == 1
            else calendar_years[0]
            if not fiscal_years and len(calendar_years) == 1
            else None
        )
        calendar_year = calendar_years[0] if len(calendar_years) == 1 else None
        period = periods[0] if len(periods) == 1 else None
        statement_duplicate_conflicts: dict[str, list[str]] = {}
        statement_duplicate_counts: dict[str, int] = {}
        statement_duplicate_mismatch = False
        for family in ("income", "cash_flow", "balance"):
            conflicts = row.get(f"{family}_statement_duplicate_conflicting_fields")
            statement_duplicate_conflicts[family] = (
                sorted(str(field) for field in conflicts)
                if isinstance(conflicts, list)
                else []
            )
            count = _safe_float(row.get(f"{family}_statement_duplicate_count"))
            statement_duplicate_counts[family] = int(count) if count is not None else 0
            mismatch = row.get(f"{family}_statement_duplicate_mismatch")
            try:
                if pd.isna(mismatch):
                    mismatch = False
            except (TypeError, ValueError):
                pass
            statement_duplicate_mismatch = statement_duplicate_mismatch or bool(mismatch)
        reported_currencies = sorted(
            {
                str(value).upper().strip()
                for key in ("income_currency", "cash_flow_currency", "balance_currency")
                if (value := row.get(key)) not in (None, "") and not pd.isna(value)
            }
        )
        currency_mismatch = len(reported_currencies) > 1
        capex_raw = _safe_float(row.get("capital_expenditures_raw"))
        capex = abs(capex_raw) if capex_raw is not None else None
        cfo = _safe_float(row.get("cash_from_operations"))
        revenue = _safe_float(row.get("revenue"))
        operating_income = _safe_float(row.get("operating_income"))
        tax_expense = _safe_float(row.get("tax_expense"))
        pretax = _safe_float(row.get("pretax_income"))
        row_tax_rate = _ratio(tax_expense, pretax)
        if row_tax_rate is None or row_tax_rate < 0 or row_tax_rate > 0.45:
            row_tax_rate = tax_rate
        short_term_debt = _safe_float(row.get("short_term_debt"))
        long_term_debt = _safe_float(row.get("long_term_debt"))
        debt = _safe_float(row.get("total_debt"))
        if debt is None:
            debt = _sum_existing(row, ["short_term_debt", "long_term_debt"])
        explicit_lease_not_in_debt = _safe_float(row.get("lease_liabilities_not_in_debt"))
        capital_lease_obligations = _safe_float(row.get("capital_lease_obligations"))
        lease_debt_reconciliation: str | None = None
        if explicit_lease_not_in_debt is not None:
            lease_liabilities_not_in_debt = explicit_lease_not_in_debt
            lease_debt_reconciliation = "explicit_not_in_debt"
        elif (
            capital_lease_obligations is not None
            and debt is not None
            and short_term_debt is not None
            and long_term_debt is not None
        ):
            debt_components = short_term_debt + long_term_debt
            tolerance = max(1.0, abs(debt), abs(debt_components + capital_lease_obligations)) * 0.02
            if abs(debt - debt_components) <= tolerance:
                lease_liabilities_not_in_debt = capital_lease_obligations
                lease_debt_reconciliation = "excluded_from_total_debt"
            elif abs(debt - (debt_components + capital_lease_obligations)) <= tolerance:
                lease_liabilities_not_in_debt = 0.0
                lease_debt_reconciliation = "included_in_total_debt"
            else:
                lease_liabilities_not_in_debt = None
                lease_debt_reconciliation = "unreconciled"
        else:
            lease_liabilities_not_in_debt = None
        equity = _safe_float(row.get("total_equity"))
        cash = _safe_float(row.get("cash"))
        short_term_investments = _safe_float(row.get("short_term_investments"))
        combined_cash_investments = _safe_float(row.get("cash_and_short_term_investments"))
        cash_investment_reconciliation_passed: bool | None = None
        if cash is None and combined_cash_investments is not None:
            cash = combined_cash_investments
            non_operating_investments = 0.0
            cash_includes_short_term_investments = True
            cash_investment_reconciliation_passed = True
        else:
            derived_investments = (
                combined_cash_investments - cash
                if combined_cash_investments is not None and cash is not None
                else None
            )
            if derived_investments is not None and derived_investments < -max(1.0, abs(cash) * 0.02):
                non_operating_investments = None
                cash_investment_reconciliation_passed = False
            elif short_term_investments is None and derived_investments is not None:
                non_operating_investments = max(0.0, derived_investments)
                cash_investment_reconciliation_passed = True
            elif short_term_investments is not None and derived_investments is not None:
                difference = abs(short_term_investments - max(0.0, derived_investments)) / max(
                    abs(short_term_investments),
                    abs(derived_investments),
                    1.0,
                )
                cash_investment_reconciliation_passed = difference <= 0.05
                non_operating_investments = short_term_investments if cash_investment_reconciliation_passed else None
            else:
                non_operating_investments = short_term_investments
            cash_includes_short_term_investments = False
        invested_capital = debt + equity - cash if debt is not None and equity is not None and cash is not None else None
        free_cash_flow = cfo - capex if cfo is not None and capex is not None else None
        nopat = operating_income * (1 - row_tax_rate) if operating_income is not None else None
        interest_expense = _safe_float(row.get("interest_expense"))
        interest_expense_sign_ambiguous = interest_expense is not None and interest_expense < 0
        fcff = (
            free_cash_flow + interest_expense * (1 - row_tax_rate)
            if free_cash_flow is not None
            and interest_expense is not None
            and not interest_expense_sign_ambiguous
            else None
        )
        stock_based_compensation = _safe_float(row.get("stock_based_compensation"))
        if stock_based_compensation is not None and stock_based_compensation < 0:
            stock_based_compensation = None
        fcff_after_sbc = (
            fcff - stock_based_compensation
            if fcff is not None and stock_based_compensation is not None
            else None
        )
        rows.append(
            {
                "date": row.get("date"),
                "fiscal_year": fiscal_year,
                "calendar_year": calendar_year,
                "period": period,
                "fiscal_years": fiscal_years,
                "calendar_years": calendar_years,
                "periods": periods,
                "fiscal_metadata_mismatch": fiscal_metadata_mismatch,
                "statement_duplicate_mismatch": statement_duplicate_mismatch,
                "statement_duplicate_conflicts": statement_duplicate_conflicts,
                "statement_duplicate_counts": statement_duplicate_counts,
                "reported_currency": reported_currencies[0] if len(reported_currencies) == 1 else None,
                "reported_currencies": reported_currencies,
                "currency_mismatch": currency_mismatch,
                "revenue": revenue,
                "gross_profit": _safe_float(row.get("gross_profit")),
                "cost_of_revenue": _safe_float(row.get("cost_of_revenue")),
                "operating_income": operating_income,
                "net_income": _safe_float(row.get("net_income")),
                "ebitda": _safe_float(row.get("ebitda")),
                "interest_expense": interest_expense,
                "interest_expense_sign_ambiguous": interest_expense_sign_ambiguous,
                "cash_from_operations": cfo,
                "capital_expenditures": capex,
                "free_cash_flow": free_cash_flow,
                "fcff": fcff,
                "fcff_after_sbc": fcff_after_sbc,
                "depreciation_amortization": _safe_float(row.get("depreciation_amortization")),
                "stock_based_compensation": stock_based_compensation,
                "common_stock_repurchased": _safe_float(row.get("common_stock_repurchased")),
                "cash": cash,
                "total_debt": debt,
                "short_term_debt": short_term_debt,
                "long_term_debt": long_term_debt,
                "total_equity": equity,
                "total_assets": _safe_float(row.get("total_assets")),
                "net_receivables": _safe_float(row.get("net_receivables")),
                "inventory": _safe_float(row.get("inventory")),
                "goodwill_and_intangibles": _safe_float(row.get("goodwill_and_intangibles")),
                "preferred_stock": _safe_float(row.get("preferred_stock")),
                "minority_interest": _safe_float(row.get("minority_interest")),
                "unfunded_pension_liability": _safe_float(row.get("unfunded_pension_liability")),
                "unfunded_pension_liability_basis": row.get("unfunded_pension_liability_basis"),
                "unfunded_pension_liability_as_of": row.get("unfunded_pension_liability_as_of"),
                "unfunded_pension_liability_source_id": row.get("unfunded_pension_liability_source_id"),
                "capital_lease_obligations": capital_lease_obligations,
                "lease_liabilities_not_in_debt": lease_liabilities_not_in_debt,
                "lease_debt_reconciliation": lease_debt_reconciliation,
                "non_operating_investments": non_operating_investments,
                "cash_includes_short_term_investments": cash_includes_short_term_investments,
                "cash_investment_reconciliation_passed": cash_investment_reconciliation_passed,
                "diluted_shares": _safe_float(row.get("diluted_shares")),
                "tax_rate": row_tax_rate,
                "nopat": nopat,
                "invested_capital": invested_capital,
                "gross_margin": _ratio(row.get("gross_profit"), revenue),
                "operating_margin": _ratio(operating_income, revenue),
                "net_margin": _ratio(row.get("net_income"), revenue),
                "fcf_margin": _ratio(free_cash_flow, revenue),
                "cash_conversion": _ratio(cfo, row.get("net_income")),
                "sbc_as_pct_revenue": _ratio(row.get("stock_based_compensation"), revenue),
                "sbc_as_pct_fcf": _ratio(row.get("stock_based_compensation"), free_cash_flow),
            }
        )

    for index, row in enumerate(rows):
        if index == 0:
            row["roic"] = None
            row["roiic"] = None
            continue
        previous = rows[index - 1]
        current_ic = _safe_float(row.get("invested_capital"))
        previous_ic = _safe_float(previous.get("invested_capital"))
        avg_ic = (current_ic + previous_ic) / 2 if current_ic is not None and previous_ic is not None else None
        row["roic"] = _ratio(row.get("nopat"), avg_ic)
        current_nopat = _safe_float(row.get("nopat"))
        previous_nopat = _safe_float(previous.get("nopat"))
        delta_nopat = current_nopat - previous_nopat if current_nopat is not None and previous_nopat is not None else None
        delta_ic = current_ic - previous_ic if current_ic is not None and previous_ic is not None else None
        row["roiic"] = _ratio(delta_nopat, delta_ic)

    return rows


def _sec_ytd_reconciliation(
    selected_rows: list[dict[str, Any]],
    company_facts: dict[str, Any] | None,
) -> dict[str, Any]:
    if not company_facts:
        return {"passed": False, "checks": [], "families": {}}

    concepts_by_metric = {
        "revenue": (SEC_CONCEPTS["income"]["revenue"], 0.02, "USD", "sum"),
        "interest_expense": (SEC_CONCEPTS["income"]["interestExpense"], 0.05, "USD", "sum"),
        "diluted_shares": (
            SEC_CONCEPTS["income"]["weightedAverageShsOutDil"],
            0.02,
            "shares",
            "weighted_average",
        ),
        "cash_from_operations": (
            SEC_CONCEPTS["cash_flow"]["netCashProvidedByOperatingActivities"],
            0.03,
            "USD",
            "sum",
        ),
        "capital_expenditures": (
            SEC_CONCEPTS["cash_flow"]["capitalExpenditure"],
            0.05,
            "USD",
            "sum",
        ),
        "stock_based_compensation": (
            SEC_CONCEPTS["cash_flow"]["stockBasedCompensation"],
            0.08,
            "USD",
            "sum",
        ),
    }
    us_gaap = _sec_us_gaap_facts(company_facts)
    fiscal_periods = ("Q1", "Q2", "Q3", "Q4")
    provider_rows = sorted(
        selected_rows,
        key=lambda row: pd.to_datetime(row.get("date"), errors="coerce"),
    )
    provider_periods = [str(row.get("period") or "").upper().strip() for row in provider_rows]
    sequence_valid = bool(
        len(provider_rows) == 4
        and all(period in fiscal_periods for period in provider_periods)
        and all(row.get("fiscal_metadata_mismatch") is not True for row in provider_rows)
        and all(
            provider_periods[index]
            == fiscal_periods[(fiscal_periods.index(provider_periods[index - 1]) + 1) % 4]
            for index in range(1, len(provider_periods))
        )
    )
    all_checks: list[dict[str, Any]] = []
    families: dict[str, dict[str, Any]] = {}

    duration_bounds = {
        "Q1": (55, 135),
        "Q2": (140, 235),
        "Q3": (220, 335),
        "Q4": (315, 430),
    }

    def normalized_fiscal_year(value: Any) -> int | None:
        number = _safe_float(value)
        if number is None or not float(number).is_integer():
            return None
        year = int(number)
        return year if 1900 <= year <= 2200 else None

    def prepared_facts(raw_facts: list[dict[str, Any]]) -> list[dict[str, Any]]:
        prepared: list[dict[str, Any]] = []
        for raw in raw_facts:
            start = pd.to_datetime(raw.get("start"), errors="coerce")
            end = pd.to_datetime(raw.get("end"), errors="coerce")
            fiscal_year = normalized_fiscal_year(raw.get("fy"))
            value = _safe_float(raw.get("val"))
            if pd.isna(start) or pd.isna(end) or end <= start or fiscal_year is None or value is None:
                continue
            prepared.append(
                {
                    **raw,
                    "_start": start,
                    "_end": end,
                    "_duration": int((end - start).days),
                    "_fiscal_year": fiscal_year,
                    "_value": float(value),
                }
            )
        return prepared

    def choose_period_fact(
        facts: list[dict[str, Any]],
        *,
        period: str,
        end_date: str,
    ) -> dict[str, Any] | None:
        fact_period = "FY" if period == "Q4" else period
        accepted_forms = SEC_ANNUAL_FORMS if period == "Q4" else {"10-Q", "10-Q/A"}
        lower_days, upper_days = duration_bounds[period]
        candidates = [
            fact
            for fact in facts
            if str(fact.get("form") or "").upper() in accepted_forms
            and str(fact.get("fp") or "").upper() == fact_period
            and str(fact.get("end") or "") == end_date
            and lower_days <= int(fact["_duration"]) <= upper_days
        ]
        if not candidates:
            return None
        return max(candidates, key=lambda fact: (int(fact["_duration"]), str(fact.get("filed") or "")))

    def choose_preceding_ytd_fact(
        facts: list[dict[str, Any]],
        *,
        period: str,
        current: dict[str, Any],
    ) -> dict[str, Any] | None:
        preceding_period = {"Q2": "Q1", "Q3": "Q2", "Q4": "Q3"}[period]
        lower_days, upper_days = duration_bounds[preceding_period]
        candidates = [
            fact
            for fact in facts
            if str(fact.get("form") or "").upper() in {"10-Q", "10-Q/A"}
            and str(fact.get("fp") or "").upper() == preceding_period
            and fact["_fiscal_year"] == current["_fiscal_year"]
            and fact["_start"] == current["_start"]
            and fact["_end"] < current["_end"]
            and 55 <= int((current["_end"] - fact["_end"]).days) <= 130
            and lower_days <= int(fact["_duration"]) <= upper_days
        ]
        if not candidates:
            return None
        return max(candidates, key=lambda fact: (fact["_end"], str(fact.get("filed") or "")))

    def derive_sec_quarters(
        raw_facts: list[dict[str, Any]],
        *,
        aggregation: str,
    ) -> tuple[list[dict[str, Any]], bool]:
        if not sequence_valid:
            return [], False
        facts = prepared_facts(raw_facts)
        derived: list[dict[str, Any]] = []
        for provider_row, period in zip(provider_rows, provider_periods, strict=True):
            parsed_end = pd.to_datetime(provider_row.get("date"), errors="coerce")
            if pd.isna(parsed_end):
                return derived, False
            end_date = str(parsed_end.date())
            current = choose_period_fact(facts, period=period, end_date=end_date)
            if current is None:
                return derived, False
            provider_fiscal_year = normalized_fiscal_year(provider_row.get("fiscal_year"))
            if provider_fiscal_year is not None and provider_fiscal_year != current["_fiscal_year"]:
                return derived, False

            if period == "Q1":
                sec_discrete = current["_value"]
                derivation = "sec_q1_duration"
                preceding_end = None
            else:
                preceding = choose_preceding_ytd_fact(facts, period=period, current=current)
                if preceding is None:
                    return derived, False
                if aggregation == "weighted_average":
                    incremental_days = int((current["_end"] - preceding["_end"]).days)
                    if incremental_days <= 0:
                        return derived, False
                    sec_discrete = (
                        current["_value"] * current["_duration"]
                        - preceding["_value"] * preceding["_duration"]
                    ) / incremental_days
                else:
                    sec_discrete = current["_value"] - preceding["_value"]
                derivation = "sec_fy_less_q3_ytd" if period == "Q4" else f"sec_{period.lower()}_ytd_less_prior_ytd"
                preceding_end = str(preceding["_end"].date())
            derived.append(
                {
                    "provider_row": provider_row,
                    "period": period,
                    "end": end_date,
                    "fiscal_year": current["_fiscal_year"],
                    "sec_discrete": sec_discrete,
                    "derivation": derivation,
                    "preceding_end": preceding_end,
                }
            )

        for index in range(1, len(derived)):
            previous = derived[index - 1]
            current = derived[index]
            expected_fiscal_year = previous["fiscal_year"] + 1 if previous["period"] == "Q4" else previous["fiscal_year"]
            if current["fiscal_year"] != expected_fiscal_year:
                return derived, False
        return derived, len(derived) == 4

    for metric, (concept_names, tolerance, unit, aggregation) in concepts_by_metric.items():
        best_derived: list[dict[str, Any]] = []
        coverage_complete = False
        selected_concept: str | None = None
        for concept in concept_names:
            payload = us_gaap.get(concept) or {}
            units = payload.get("units") if isinstance(payload, dict) else {}
            candidates = units.get(unit) if isinstance(units, dict) else None
            facts = [fact for fact in (candidates or []) if isinstance(fact, dict)]
            if not facts:
                continue
            derived, complete = derive_sec_quarters(facts, aggregation=aggregation)
            if complete or len(derived) > len(best_derived):
                best_derived = derived
                selected_concept = concept
                coverage_complete = complete
            if complete:
                break

        metric_checks: list[dict[str, Any]] = []
        for quarter in best_derived:
            provider_value = _safe_float(quarter["provider_row"].get(metric))
            if provider_value is None:
                continue
            sec_discrete = float(quarter["sec_discrete"])
            if metric == "capital_expenditures":
                sec_discrete = abs(sec_discrete)
            difference = abs(provider_value - sec_discrete) / max(abs(provider_value), abs(sec_discrete), 1.0)
            metric_checks.append(
                {
                    "metric": metric,
                    "period": quarter["period"],
                    "end": quarter["end"],
                    "fiscal_year": quarter["fiscal_year"],
                    "concept": selected_concept,
                    "provider_discrete": provider_value,
                    "sec_discrete": sec_discrete,
                    "derivation": quarter["derivation"],
                    "preceding_end": quarter["preceding_end"],
                    "difference": difference,
                    "maximum_difference": tolerance,
                    "passed": difference <= tolerance,
                }
            )

        family_passed = bool(
            sequence_valid
            and coverage_complete
            and len(metric_checks) == 4
            and all(check["passed"] for check in metric_checks)
        )
        families[metric] = {
            "passed": family_passed,
            "concept": selected_concept,
            "unit": unit,
            "aggregation": aggregation,
            "sequence_valid": sequence_valid,
            "coverage_complete": coverage_complete,
            "expected_quarters": list(provider_periods),
            "checked_quarters": [check["period"] for check in metric_checks],
            "checks": metric_checks,
        }
        all_checks.extend(metric_checks)

    return {
        "passed": bool(families) and all(item["passed"] for item in families.values()),
        "checks": all_checks,
        "families": families,
    }


def _build_ttm_row(
    frames: dict[str, Any],
    tax_rate: float,
    *,
    sec_company_facts: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    quarterly_frames = {
        "income": frames.get("income_quarterly", pd.DataFrame()),
        "cash_flow": frames.get("cash_flow_quarterly", pd.DataFrame()),
        "balance": frames.get("balance_quarterly", pd.DataFrame()),
    }
    if any(not isinstance(frame, pd.DataFrame) or frame.empty for frame in quarterly_frames.values()):
        return None
    def frame_quarters(frame: pd.DataFrame) -> dict[str, str | None]:
        output: dict[str, str | None] = {}
        if frame.empty or "date" not in frame.columns:
            return output
        for _, raw in frame.iterrows():
            parsed = pd.to_datetime(raw.get("date"), errors="coerce")
            if pd.isna(parsed):
                continue
            date = str(parsed.date())
            period = str(raw.get("period") or raw.get("fiscalPeriod") or "").upper().strip() or None
            output[date] = period
        return output

    quarter_maps = {key: frame_quarters(frame) for key, frame in quarterly_frames.items()}
    common_dates = set.intersection(*(set(values) for values in quarter_maps.values()))
    quarterly_rows = _normalize_financials(quarterly_frames, tax_rate)
    complete = [
        row
        for row in quarterly_rows
        if row.get("date") in common_dates
        if _safe_float(row.get("revenue")) is not None
        and _safe_float(row.get("cash_from_operations")) is not None
        and _safe_float(row.get("capital_expenditures")) is not None
    ]
    if len(complete) < 4:
        return None
    selected = complete[-4:]
    dates = [pd.to_datetime(row.get("date"), errors="coerce") for row in selected]
    if any(pd.isna(value) for value in dates) or len({str(value.date()) for value in dates}) != 4:
        return None
    dates = sorted(dates)
    gaps = [(dates[index] - dates[index - 1]).days for index in range(1, len(dates))]
    span_days = (dates[-1] - dates[0]).days
    if span_days < 240 or span_days > 380 or any(gap < 55 or gap > 130 for gap in gaps):
        return None
    selected_dates = {str(value.date()) for value in dates}
    known_periods = [
        period
        for mapping in quarter_maps.values()
        for date, period in mapping.items()
        if date in selected_dates and period is not None
    ]
    if any(period not in {"Q1", "Q2", "Q3", "Q4"} for period in known_periods):
        return None
    discrete_periods_confirmed = len(known_periods) >= 8

    def sum_field(key: str) -> float | None:
        values = [_safe_float(row.get(key)) for row in selected]
        return sum(value for value in values if value is not None) if all(value is not None for value in values) else None

    latest = selected[-1]
    shares = [_safe_float(row.get("diluted_shares")) for row in selected]
    shares = [value for value in shares if value is not None and value > 0]
    diluted_share_quarters = len(shares)
    operating_income = sum_field("operating_income")
    tax_rates = [_safe_float(row.get("tax_rate")) for row in selected]
    tax_rates = [value for value in tax_rates if value is not None and 0 <= value <= 0.45]
    normalized_tax_rate = float(pd.Series(tax_rates).median()) if tax_rates else tax_rate
    free_cash_flow = sum_field("free_cash_flow")
    interest_expense = sum_field("interest_expense")
    interest_expense_sign_ambiguous = any(row.get("interest_expense_sign_ambiguous") is True for row in selected)
    fcff = (
        free_cash_flow + interest_expense * (1 - normalized_tax_rate)
        if free_cash_flow is not None
        and interest_expense is not None
        and not interest_expense_sign_ambiguous
        else None
    )
    stock_based_compensation = sum_field("stock_based_compensation")
    fcff_after_sbc = (
        fcff - stock_based_compensation
        if fcff is not None and stock_based_compensation is not None and stock_based_compensation >= 0
        else None
    )
    revenue = sum_field("revenue")
    net_income = sum_field("net_income")
    cfo = sum_field("cash_from_operations")
    ttm_currencies = sorted(
        {
            str(currency).upper().strip()
            for row in selected
            for currency in (row.get("reported_currencies") or ([row.get("reported_currency")] if row.get("reported_currency") else []))
            if currency
        }
    )
    ttm_currency_mismatch = len(ttm_currencies) > 1 or any(row.get("currency_mismatch") is True for row in selected)
    result = {
        "date": str(max(dates).date()),
        "period": "TTM",
        "source": "fmp_quarterly_ttm",
        "reported_currency": ttm_currencies[0] if len(ttm_currencies) == 1 else None,
        "reported_currencies": ttm_currencies,
        "currency_mismatch": ttm_currency_mismatch,
        "ttm_validation": {
            "status": "date_sequence_only",
            "quarter_dates": sorted(selected_dates),
            "span_days": span_days,
            "gaps_days": gaps,
            "discrete_periods_confirmed": discrete_periods_confirmed,
            "provider_ttm_reconciled": False,
            "provider_ttm_checks": [],
            "diluted_share_quarters": diluted_share_quarters,
        },
        "revenue": revenue,
        "gross_profit": sum_field("gross_profit"),
        "cost_of_revenue": sum_field("cost_of_revenue"),
        "operating_income": operating_income,
        "net_income": net_income,
        "ebitda": sum_field("ebitda"),
        "interest_expense": interest_expense,
        "interest_expense_sign_ambiguous": interest_expense_sign_ambiguous,
        "cash_from_operations": cfo,
        "capital_expenditures": sum_field("capital_expenditures"),
        "free_cash_flow": free_cash_flow,
        "fcff": fcff,
        "fcff_after_sbc": fcff_after_sbc,
        "depreciation_amortization": sum_field("depreciation_amortization"),
        "stock_based_compensation": stock_based_compensation,
        "common_stock_repurchased": sum_field("common_stock_repurchased"),
        "cash": _safe_float(latest.get("cash")),
        "total_debt": _safe_float(latest.get("total_debt")),
        "short_term_debt": _safe_float(latest.get("short_term_debt")),
        "long_term_debt": _safe_float(latest.get("long_term_debt")),
        "total_equity": _safe_float(latest.get("total_equity")),
        "total_assets": _safe_float(latest.get("total_assets")),
        "net_receivables": _safe_float(latest.get("net_receivables")),
        "inventory": _safe_float(latest.get("inventory")),
        "goodwill_and_intangibles": _safe_float(latest.get("goodwill_and_intangibles")),
        "preferred_stock": _safe_float(latest.get("preferred_stock")),
        "minority_interest": _safe_float(latest.get("minority_interest")),
        "unfunded_pension_liability": _safe_float(latest.get("unfunded_pension_liability")),
        "unfunded_pension_liability_basis": latest.get("unfunded_pension_liability_basis"),
        "unfunded_pension_liability_as_of": latest.get("unfunded_pension_liability_as_of"),
        "unfunded_pension_liability_source_id": latest.get("unfunded_pension_liability_source_id"),
        "capital_lease_obligations": _safe_float(latest.get("capital_lease_obligations")),
        "lease_liabilities_not_in_debt": _safe_float(latest.get("lease_liabilities_not_in_debt")),
        "lease_debt_reconciliation": latest.get("lease_debt_reconciliation"),
        "non_operating_investments": _safe_float(latest.get("non_operating_investments")),
        "cash_includes_short_term_investments": latest.get("cash_includes_short_term_investments") is True,
        "cash_investment_reconciliation_passed": latest.get("cash_investment_reconciliation_passed"),
        "diluted_shares": float(pd.Series(shares).mean()) if diluted_share_quarters == 4 else None,
        "tax_rate": normalized_tax_rate,
        "nopat": operating_income * (1 - normalized_tax_rate) if operating_income is not None else None,
        "invested_capital": _safe_float(latest.get("invested_capital")),
        "gross_margin": _ratio(sum_field("gross_profit"), revenue),
        "operating_margin": _ratio(operating_income, revenue),
        "net_margin": _ratio(net_income, revenue),
        "fcf_margin": _ratio(free_cash_flow, revenue),
        "cash_conversion": _ratio(cfo, net_income),
        "sbc_as_pct_revenue": _ratio(sum_field("stock_based_compensation"), revenue),
        "sbc_as_pct_fcf": _ratio(sum_field("stock_based_compensation"), free_cash_flow),
        "roic": _ratio(operating_income * (1 - normalized_tax_rate) if operating_income is not None else None, _safe_float(latest.get("invested_capital"))),
        "roiic": None,
    }

    provider_frames = {
        "income": frames.get("income_ttm", pd.DataFrame()),
        "cash_flow": frames.get("cash_flow_ttm", pd.DataFrame()),
        "balance": frames.get("balance_ttm", pd.DataFrame()),
    }
    provider_dates: dict[str, str] = {}
    aligned_provider_frames: dict[str, pd.DataFrame] = {}
    for statement_key, frame in provider_frames.items():
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            aligned_provider_frames[statement_key] = pd.DataFrame()
            continue
        candidate = frame.copy()
        if "date" in candidate.columns:
            candidate["date"] = pd.to_datetime(candidate["date"], errors="coerce")
            candidate = candidate.sort_values("date", na_position="first")
            source_date = candidate.iloc[-1].get("date")
            if not pd.isna(source_date):
                provider_dates[statement_key] = str(source_date.date())
        candidate = candidate.iloc[[-1]].copy()
        candidate["date"] = result["date"]
        aligned_provider_frames[statement_key] = candidate

    provider_rows = _normalize_financials(aligned_provider_frames, tax_rate)
    provider_row = provider_rows[-1] if provider_rows else None
    balance_provider_rows = _normalize_financials(
        {
            "income": pd.DataFrame(),
            "cash_flow": pd.DataFrame(),
            "balance": aligned_provider_frames["balance"],
        },
        tax_rate,
    ) if "balance" in aligned_provider_frames else []
    balance_provider_row = balance_provider_rows[-1] if balance_provider_rows else None
    tolerances = {
        "revenue": 0.02,
        "net_income": 0.04,
        "ebitda": 0.04,
        "interest_expense": 0.05,
        "cash_from_operations": 0.03,
        "capital_expenditures": 0.05,
        "free_cash_flow": 0.05,
        "stock_based_compensation": 0.08,
        "diluted_shares": 0.02,
        "cash": 0.03,
        "total_debt": 0.03,
        "total_equity": 0.03,
        "preferred_stock": 0.05,
        "minority_interest": 0.05,
        "lease_liabilities_not_in_debt": 0.05,
        "non_operating_investments": 0.05,
    }
    provider_checks: list[dict[str, Any]] = []
    if provider_row:
        for metric, tolerance in tolerances.items():
            calculated_value = _safe_float(result.get(metric))
            provider_value = _safe_float(provider_row.get(metric))
            if calculated_value is None or provider_value is None:
                continue
            difference = abs(calculated_value - provider_value) / max(abs(calculated_value), abs(provider_value), 1.0)
            provider_checks.append(
                {
                    "metric": metric,
                    "calculated_value": calculated_value,
                    "provider_value": provider_value,
                    "difference": difference,
                    "maximum_difference": tolerance,
                    "passed": difference <= tolerance,
                }
            )

    required_metrics = {
        "revenue",
        "cash_from_operations",
        "capital_expenditures",
        "interest_expense",
        "stock_based_compensation",
        "diluted_shares",
    }
    required_checks = [check for check in provider_checks if check["metric"] in required_metrics]
    latest_quarter_date = pd.to_datetime(result["date"])
    provider_date_differences = {
        key: int((latest_quarter_date - pd.to_datetime(value)).days)
        for key, value in provider_dates.items()
    }
    provider_dates_current = bool(provider_date_differences) and all(
        0 <= days <= 45 for days in provider_date_differences.values()
    )
    balance_date_gap_days = provider_date_differences.get("balance")
    provider_balance_date_current = bool(
        balance_date_gap_days is not None
        and 0 <= balance_date_gap_days <= 45
    )
    provider_currency = str((provider_row or {}).get("reported_currency") or "").upper().strip()
    balance_provider_currency = str((balance_provider_row or {}).get("reported_currency") or "").upper().strip()
    calculated_currency = str(result.get("reported_currency") or "").upper().strip()
    currency_reconciled = bool(
        provider_currency
        and calculated_currency
        and provider_currency == calculated_currency
    )
    provider_balance_currency_reconciled = bool(
        balance_provider_currency
        and calculated_currency
        and balance_provider_currency == calculated_currency
    )
    provider_ttm_reconciled = bool(
        discrete_periods_confirmed
        and len(required_checks) == len(required_metrics)
        and provider_dates_current
        and currency_reconciled
        and provider_balance_date_current
        and provider_balance_currency_reconciled
        and all(check["passed"] for check in provider_checks)
    )
    sec_ttm_reconciliation = _sec_ytd_reconciliation(selected, sec_company_facts)
    sec_quarterly_reconciled = sec_ttm_reconciliation["passed"] is True
    provider_reconciled_metrics = sorted(
        check["metric"] for check in provider_checks if check.get("passed") is True
    )
    sec_reconciled_metrics = sorted(
        metric
        for metric, family in sec_ttm_reconciliation["families"].items()
        if family.get("passed") is True
    )
    provider_available = bool(provider_row and required_checks)
    validation_status = (
        "validated"
        if provider_ttm_reconciled or sec_quarterly_reconciled
        else "provider_ttm_mismatch"
        if provider_available
        else "date_sequence_only"
    )
    result["ttm_validation"].update(
        {
            "status": validation_status,
            "period_basis": (
                "provider_ttm_reconciled"
                if provider_ttm_reconciled
                else "discrete_reconciled_to_sec_ytd"
                if sec_quarterly_reconciled
                else "unverified"
            ),
            "provider_ttm_reconciled": provider_ttm_reconciled,
            "provider_reconciled_metrics": provider_reconciled_metrics,
            "provider_ttm_checks": provider_checks,
            "provider_ttm_dates": provider_dates,
            "provider_ttm_date_gaps_days": provider_date_differences,
            "provider_ttm_dates_current": provider_dates_current,
            "provider_ttm_balance_date": provider_dates.get("balance"),
            "provider_ttm_balance_date_gap_days": balance_date_gap_days,
            "provider_ttm_balance_date_current": provider_balance_date_current,
            "provider_ttm_currency": provider_currency or None,
            "provider_ttm_balance_currency": balance_provider_currency or None,
            "provider_ttm_balance_currency_reconciled": provider_balance_currency_reconciled,
            "calculated_currency": calculated_currency or None,
            "currency_reconciled": currency_reconciled,
            "sec_quarterly_reconciled": sec_quarterly_reconciled,
            "sec_reconciled_metrics": sec_reconciled_metrics,
            "sec_quarterly_checks": sec_ttm_reconciliation["checks"],
            "sec_quarterly_families": sec_ttm_reconciliation["families"],
        }
    )
    if provider_ttm_reconciled:
        result["source"] = "fmp_quarterly_ttm_reconciled_to_fmp_statement_ttm"
    elif sec_quarterly_reconciled:
        result["source"] = "fmp_quarterly_ttm_reconciled_to_sec_ytd"
    return result


def _latest(rows: list[dict[str, Any]], key: str) -> float | None:
    for row in reversed(rows):
        value = _safe_float(row.get(key))
        if value is not None:
            return value
    return None


def _growth_between(rows: list[dict[str, Any]], key: str, years: int) -> float | None:
    clean = [row for row in rows if _safe_float(row.get(key)) is not None and _safe_float(row.get(key)) > 0]
    if len(clean) < 2:
        return None
    lookback = min(years, len(clean) - 1)
    start = _safe_float(clean[-lookback - 1].get(key))
    end = _safe_float(clean[-1].get(key))
    return calculate_revenue_cagr(start or 0, end or 0, lookback)


def _derive_assumptions(rows: list[dict[str, Any]], ttm_row: dict[str, Any] | None = None) -> dict[str, Any]:
    tax_rates = [_safe_float(row.get("tax_rate")) for row in rows]
    tax_rates = [value for value in tax_rates if value is not None and 0 <= value <= 0.45]
    fcf_margins = [_safe_float(row.get("fcf_margin")) for row in rows]
    fcf_margins = [value for value in fcf_margins if value is not None and value > 0]
    latest_fcf_margin = _safe_float((ttm_row or {}).get("fcf_margin")) or _latest(rows, "fcf_margin")
    base_margin = latest_fcf_margin if latest_fcf_margin is not None and latest_fcf_margin > 0 else (pd.Series(fcf_margins).median() if fcf_margins else None)
    revenue_growth = _growth_between(rows, "revenue", 5)

    return {
        "normalized_tax_rate": float(pd.Series(tax_rates).median()) if tax_rates else DEFAULT_TAX_RATE,
        "base_revenue_growth": float(max(min(revenue_growth, 0.45), -0.20)) if revenue_growth is not None else None,
        "base_fcf_margin": float(max(min(base_margin, 0.55), 0.0)) if base_margin is not None else None,
        "wacc": DEFAULT_WACC,
        "terminal_growth": DEFAULT_TERMINAL_GROWTH,
        "forecast_years": 5,
        "source": "historical_median_or_default_when_data_missing",
    }


def _build_ratios(rows: list[dict[str, Any]], ttm_row: dict[str, Any] | None = None) -> dict[str, Any]:
    current = ttm_row or {}
    latest_revenue = _safe_float(current.get("revenue")) or _latest(rows, "revenue")
    latest_debt = _safe_float(current.get("total_debt")) if ttm_row else _latest(rows, "total_debt")
    latest_cash = _safe_float(current.get("cash")) if ttm_row else _latest(rows, "cash")
    latest_debt = latest_debt or 0.0
    latest_cash = latest_cash or 0.0
    latest_ebitda = _safe_float(current.get("ebitda")) or _latest(rows, "ebitda")
    latest_shares = _safe_float(current.get("diluted_shares")) or _latest(rows, "diluted_shares")
    first_shares = next((_safe_float(row.get("diluted_shares")) for row in rows if _safe_float(row.get("diluted_shares")) is not None), None)
    share_years = max(len([row for row in rows if _safe_float(row.get("diluted_shares")) is not None]) - 1, 0)

    return {
        "revenue_cagr_3y": _growth_between(rows, "revenue", 3),
        "revenue_cagr_5y": _growth_between(rows, "revenue", 5),
        "gross_margin": _safe_float(current.get("gross_margin")) or _latest(rows, "gross_margin"),
        "operating_margin": _safe_float(current.get("operating_margin")) or _latest(rows, "operating_margin"),
        "net_margin": _safe_float(current.get("net_margin")) or _latest(rows, "net_margin"),
        "fcf_margin": _safe_float(current.get("fcf_margin")) or _latest(rows, "fcf_margin"),
        "roic": _safe_float(current.get("roic")) or _latest(rows, "roic"),
        "roiic": _latest(rows, "roiic"),
        "cash_conversion": _safe_float(current.get("cash_conversion")) or _latest(rows, "cash_conversion"),
        "net_debt": latest_debt - latest_cash,
        "net_debt_to_ebitda": _ratio(latest_debt - latest_cash, latest_ebitda),
        "sbc_as_pct_revenue": _safe_float(current.get("sbc_as_pct_revenue")) or _latest(rows, "sbc_as_pct_revenue"),
        "sbc_as_pct_fcf": _safe_float(current.get("sbc_as_pct_fcf")) or _latest(rows, "sbc_as_pct_fcf"),
        "share_count_cagr": calculate_revenue_cagr(first_shares or 0, latest_shares or 0, share_years) if share_years else None,
        "latest_revenue": latest_revenue,
        "latest_fcf": _safe_float(current.get("free_cash_flow")) or _latest(rows, "free_cash_flow"),
        "latest_diluted_shares": latest_shares,
        "data_period": "TTM" if ttm_row else "annual",
        "data_as_of": (ttm_row or {}).get("date") or (rows[-1].get("date") if rows else None),
    }


def _quality_flags(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    flags: list[dict[str, Any]] = []
    latest = rows[-1] if rows else {}
    revenue_growth = _growth_between(rows, "revenue", 3)
    receivables_growth = _growth_between(rows, "net_receivables", 3)
    inventory_growth = _growth_between(rows, "inventory", 3)
    cogs_growth = _growth_between(rows, "cost_of_revenue", 3)
    share_growth = _growth_between(rows, "diluted_shares", 3)

    if receivables_growth is not None and revenue_growth is not None and receivables_growth > revenue_growth + 0.05:
        flags.append({"severity": "medium", "title": "Receivables growing faster than revenue", "metric": receivables_growth})
    if inventory_growth is not None and cogs_growth is not None and inventory_growth > cogs_growth + 0.05:
        flags.append({"severity": "medium", "title": "Inventory growing faster than COGS", "metric": inventory_growth})
    if (_safe_float(latest.get("cash_conversion")) or 0) < 0.8 and (_safe_float(latest.get("net_income")) or 0) > 0:
        flags.append({"severity": "high", "title": "Cash conversion below earnings quality threshold", "metric": latest.get("cash_conversion")})
    goodwill_intangibles = _ratio(latest.get("goodwill_and_intangibles"), latest.get("total_assets"))
    if goodwill_intangibles is not None and goodwill_intangibles > 0.4:
        flags.append({"severity": "medium", "title": "High goodwill and intangible asset intensity", "metric": goodwill_intangibles})
    if share_growth is not None and share_growth > 0.02:
        flags.append({"severity": "medium", "title": "Share count creep", "metric": share_growth})
    if (_safe_float(latest.get("free_cash_flow")) or 0) < 0 and (_safe_float(latest.get("net_income")) or 0) > 0:
        flags.append({"severity": "high", "title": "Negative FCF despite positive earnings", "metric": latest.get("free_cash_flow")})
    return flags


def _build_valuation(
    ticker: str,
    rows: list[dict[str, Any]],
    ttm_row: dict[str, Any] | None,
    profile: dict[str, Any],
    frames: dict[str, Any],
    assumptions: dict[str, Any],
    source_records: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    return build_institutional_valuation(
        annual_rows=rows,
        ttm_row=ttm_row,
        profile=profile,
        quote=frames.get("quote") if isinstance(frames.get("quote"), dict) else {},
        prices=frames.get("prices", pd.DataFrame()),
        analyst_estimates=frames.get("analyst_estimates", pd.DataFrame()),
        key_metrics_ttm=frames.get("key_metrics_ttm") if isinstance(frames.get("key_metrics_ttm"), dict) else {},
        ratios_ttm=frames.get("ratios_ttm") if isinstance(frames.get("ratios_ttm"), dict) else {},
        assumptions=assumptions,
        expected_ticker=ticker,
        source_records=source_records,
    )


def _audit_bundle(
    ticker: str,
    rows: list[dict[str, Any]],
    sources: list[dict[str, Any]],
    valuation: dict[str, Any],
    data_points: list[dict[str, Any]],
) -> dict[str, Any]:
    findings: list[dict[str, Any]] = []
    coverage = _build_evidence_coverage(sources, data_points)
    if not rows:
        findings.append({"severity": "high", "code": "missing_financials", "message": f"No normalized financial statements are available for {ticker}."})
    if any(source.get("status") == "error" for source in sources):
        findings.append({"severity": "medium", "code": "provider_error", "message": "At least one provider call failed. Inspect sources.json."})
    required_fmp_ids = {"fmp:profile", "fmp:income:annual", "fmp:cash-flow:annual", "fmp:balance:annual"}
    required_fmp = [source for source in sources if source.get("source_id") in required_fmp_ids]
    fallback_source_ids = set(YFINANCE_STATEMENT_SOURCE_IDS.values()) | set(SEC_STATEMENT_SOURCE_IDS.values()) | {
        "yfinance:profile",
        "yfinance:quote",
        "yfinance:prices",
    }
    fallback_available = any(
        source.get("source_id") in fallback_source_ids
        and source.get("status") == "ok"
        and int(source.get("row_count") or 0) > 0
        for source in sources
    )
    if required_fmp and all(source.get("status") == "unavailable" for source in required_fmp) and not fallback_available:
        findings.append({"severity": "high", "code": "provider_unavailable", "message": "FMP is unavailable in this runtime, so no source-backed report can be completed."})
    if any(
        source.get("status") == "unavailable"
        and source.get("provider") == "sec-edgar"
        and str(source.get("endpoint_or_filing") or "").startswith("env:")
        for source in sources
    ):
        findings.append({"severity": "medium", "code": "sec_edgar_unavailable", "message": "SEC EDGAR metadata is unavailable because SEC_USER_AGENT is not configured."})
    reconciliation_source = next((source for source in sources if source.get("source_id") == SEC_RECONCILIATION_SOURCE_ID), {})
    reconciliation_status = reconciliation_source.get("reconciliation_status")
    if rows and reconciliation_status == "mismatch":
        findings.append({"severity": "high", "code": "filing_reconciliation_mismatch", "message": "Normalized statements do not reconcile to SEC/XBRL within the required tolerances."})
    elif rows and reconciliation_status not in {"reconciled", "primary_sec", "not_required"}:
        findings.append({"severity": "medium", "code": "filing_reconciliation_missing", "message": "SEC/XBRL facts were not numerically reconciled to the normalized statements in this run."})
    if not valuation.get("available"):
        findings.append({"severity": "medium", "code": "valuation_unavailable", "message": valuation.get("reason", "Valuation could not be completed.")})
    valuation_status = str(valuation.get("status") or "legacy")
    reliability = valuation.get("reliability") or {}
    if valuation_status == "research_grade":
        findings.append(
            {
                "severity": "medium",
                "code": "valuation_research_grade",
                "message": "The evidence supports a research range, but unresolved controls still withhold a decision-ready central value.",
            }
        )
    elif valuation_status != "decision_ready":
        findings.append(
            {
                "severity": "high",
                "code": "valuation_not_decision_ready",
                "message": "The valuation does not pass current-data, model-fit, and reliability gates; do not use it as fair value.",
            }
        )
    if (valuation.get("price_validation") or {}).get("status") == "blocked":
        findings.append({"severity": "high", "code": "price_validation_blocked", "message": "Price, share count, or market capitalization could not be reconciled."})
    share_reconciliation = valuation.get("share_denominator_reconciliation") or {}
    if share_reconciliation.get("passed") is False:
        findings.append(
            {
                "severity": "high",
                "code": "share_denominator_mismatch",
                "message": "Current basic outstanding shares do not reconcile to TTM diluted weighted-average shares within the control tolerance.",
            }
        )
    terminal_share = _safe_float(reliability.get("terminal_value_share"))
    if terminal_share is not None and terminal_share > 0.85:
        findings.append({"severity": "high", "code": "terminal_value_dominates", "message": "More than 85% of modeled value comes from the terminal period."})
    reverse_status = str((valuation.get("reverse_dcf") or {}).get("status") or "")
    if reverse_status in {"above_range", "below_range"}:
        findings.append({"severity": "medium", "code": "reverse_dcf_outside_range", "message": "Reverse DCF is outside the solved range and must be shown as a bound, not an exact growth rate."})
    for point in data_points:
        if point["claim_tag"] == "sourced_fact" and not point.get("source_id"):
            findings.append({"severity": "high", "code": "missing_source", "message": f"{point['metric']} lacks a source id."})
    if coverage["score"] < 85:
        severity = "high" if coverage["score"] < 60 else "medium"
        findings.append(
            {
                "severity": severity,
                "code": "evidence_coverage_gap",
                "message": f"Evidence ledger covers {coverage['covered_expected_metrics']}/{coverage['expected_metrics']} required research metrics.",
            }
        )
    if coverage["sourced_points_missing_ok_source"]:
        findings.append(
            {
                "severity": "high",
                "code": "sourced_point_without_ok_source",
                "message": f"{len(coverage['sourced_points_missing_ok_source'])} sourced facts do not map to an ok source record.",
            }
        )
    if coverage["calculated_points_missing_formula"]:
        findings.append(
            {
                "severity": "medium",
                "code": "formula_missing",
                "message": f"{len(coverage['calculated_points_missing_formula'])} calculated metrics are missing formulas.",
            }
        )
    return {
        "generated_at": _now_iso(),
        "status": "pass" if not [item for item in findings if item["severity"] == "high"] else "needs_attention",
        "findings": findings,
        "coverage": coverage,
    }


def _checklist_scores(ratios: dict[str, Any], quality_flags: list[dict[str, Any]], valuation: dict[str, Any], audit: dict[str, Any] | None = None) -> dict[str, Any]:
    roic = ratios.get("roic") or 0
    fcf_margin = ratios.get("fcf_margin") or 0
    cash_conversion = ratios.get("cash_conversion") or 0
    reliability_score = _safe_float((valuation.get("reliability") or {}).get("score")) or 0.0
    if valuation.get("status") == "not_decision_ready":
        reliability_score = min(reliability_score, 0.49)
    return {
        "quality": round(max(0, min(100, 35 + roic * 180 + fcf_margin * 80 + cash_conversion * 15))),
        "accounting_risk": round(max(0, 100 - len(quality_flags) * 18)),
        "valuation": round(max(0, min(100, reliability_score * 100))),
        "evidence": round(max(0, min(100, (audit or {}).get("coverage", {}).get("score", 40)))),
    }


def _valuation_precision_is_backed(valuation: dict[str, Any], audit: dict[str, Any]) -> bool:
    reliability = valuation.get("reliability") or {}
    price_validation = valuation.get("price_validation") or {}
    return bool(
        valuation.get("model_version") == "institutional_valuation_v3"
        and valuation.get("available")
        and valuation.get("status") == "decision_ready"
        and reliability.get("usable") is True
        and reliability.get("status") == "high"
        and price_validation.get("usable") is True
        and price_validation.get("status") == "validated"
        and audit.get("status") == "pass"
    )


def _valuation_for_output(valuation: dict[str, Any], audit: dict[str, Any]) -> dict[str, Any]:
    output = deepcopy(valuation)
    backed = _valuation_precision_is_backed(valuation, audit)
    output["precision_withheld"] = not backed
    if backed:
        return output

    reliability = valuation.get("reliability") or {}
    readiness_gates = reliability.get("readiness_gates") or {}
    research_range_visible = bool(
        valuation.get("status") == "research_grade"
        and reliability.get("usable") is True
        and (readiness_gates.get("fresh_market_data") or {}).get("passed") is True
        and (readiness_gates.get("fresh_financial_data") or {}).get("passed") is True
        and (readiness_gates.get("currency_consistency") or {}).get("passed") is True
        and (readiness_gates.get("fundamental_scale_reconciliation") or {}).get("passed") is True
    )
    raw_range = valuation.get("range") or {}
    output["range"] = {
        "low": raw_range.get("low") if research_range_visible else None,
        "central": None,
        "high": raw_range.get("high") if research_range_visible else None,
    }
    output["selected_value"] = None
    output["scenarios"] = [
        {
            "name": scenario.get("name"),
            "method": scenario.get("method"),
            "assumptions": scenario.get("assumptions") or {},
            "terminal_value_share": scenario.get("terminal_value_share"),
            "intrinsic_value_per_share": None,
            "equity_value": None,
            "enterprise_value": None,
            "terminal_value": None,
            "pv_terminal_value": None,
            "pv_explicit_cash_flow": None,
            "forecast": [],
        }
        for scenario in valuation.get("scenarios", [])
    ]
    output["methods"] = [
        {
            "key": method.get("key"),
            "role": method.get("role"),
            "weight": method.get("weight"),
            "currency": method.get("currency"),
            "value_per_share": None,
        }
        for method in valuation.get("methods", [])
    ]
    reverse = valuation.get("reverse_dcf") or {}
    output["reverse_dcf"] = {
        "available": reverse.get("available", False),
        "status": reverse.get("status"),
        "reason": reverse.get("reason"),
        "bound": reverse.get("bound"),
        "weight": 0,
        "implied_revenue_cagr": None,
        "current_price": None,
        "value_at_floor": None,
        "value_at_ceiling": None,
    }
    return output


def _data_points_for_output(data_points: list[dict[str, Any]], *, backed: bool) -> list[dict[str, Any]]:
    if backed:
        return data_points
    sensitive_fragments = (
        "intrinsic_value_per_share",
        ".equity_value",
        ".terminal_value",
        ".pv_terminal_value",
        "valuation.range.central",
        "valuation_range_central",
        "valuation.reverse_dcf.implied_revenue_cagr",
        "valuation_reverse_dcf_implied_revenue_cagr",
        "reverse_dcf_implied_revenue_cagr",
        "valuation.reverse_dcf.value_at_floor",
        "valuation.reverse_dcf.value_at_ceiling",
    )
    output: list[dict[str, Any]] = []
    for point in data_points:
        metric = str(point.get("metric") or "")
        if any(fragment in metric for fragment in sensitive_fragments):
            output.append(
                {
                    **point,
                    "raw_value": None,
                    "normalized_value": None,
                    "claim_tag": "uncertainty",
                    "formula": "exact value withheld until all valuation reliability gates pass",
                }
            )
        else:
            output.append(point)
    return output


def _canonical_assumptions(valuation: dict[str, Any], fallback: dict[str, Any]) -> dict[str, Any]:
    base = next((item for item in valuation.get("scenarios", []) if item.get("name") == "base"), {})
    scenario_assumptions = base.get("assumptions") or {}
    capital = valuation.get("cost_of_capital") or {}
    return {
        "model_version": valuation.get("model_version"),
        "valuation_status": valuation.get("status"),
        "archetype": valuation.get("archetype"),
        "primary_method": valuation.get("primary_method"),
        "cash_flow_basis": valuation.get("cash_flow_basis"),
        "market_data_as_of": valuation.get("market_data_as_of"),
        "financial_data_as_of": valuation.get("financial_data_as_of"),
        "risk_free_rate": capital.get("risk_free_rate"),
        "equity_risk_premium": capital.get("equity_risk_premium"),
        "raw_beta": capital.get("raw_beta"),
        "adjusted_beta": capital.get("adjusted_beta", capital.get("beta")),
        "cost_of_equity": capital.get("cost_of_equity"),
        "operating_beta": capital.get("operating_beta"),
        "operating_discount_rate": capital.get("operating_discount_rate"),
        "tax_shield_present_value": capital.get("tax_shield_present_value"),
        "tax_shield_upper_bound": capital.get("tax_shield_upper_bound"),
        "wacc": capital.get("wacc") if valuation.get("primary_method") in {"forward_fcff_dcf", "through_cycle_fcff_dcf"} else None,
        "discount_rate": scenario_assumptions.get("discount_rate") or scenario_assumptions.get("wacc") or scenario_assumptions.get("cost_of_equity"),
        "terminal_growth": scenario_assumptions.get("terminal_growth"),
        "forecast_years": scenario_assumptions.get("years") or fallback.get("forecast_years"),
        "historical_revenue_growth_reference": fallback.get("base_revenue_growth"),
        "historical_cash_flow_margin_reference": fallback.get("base_fcf_margin"),
        "consensus_role": (valuation.get("model_policy") or {}).get("consensus_role"),
    }


def _assumptions_yml(assumptions: dict[str, Any]) -> str:
    lines = ["assumptions:"]
    for key, value in assumptions.items():
        lines.append(f"  {key}: {value}")
    return "\n".join(lines) + "\n"


def _json_text(payload: Any) -> str:
    return json.dumps(payload, indent=2, sort_keys=True, default=str)


def _text_download(filename: str, media_type: str, text: str) -> dict[str, Any]:
    return {
        "filename": filename,
        "media_type": media_type,
        "encoding": "base64",
        "content_base64": base64.b64encode(text.encode("utf-8")).decode("ascii"),
    }


def _bytes_download(filename: str, media_type: str, content: bytes) -> dict[str, Any]:
    return {
        "filename": filename,
        "media_type": media_type,
        "encoding": "base64",
        "content_base64": base64.b64encode(content).decode("ascii"),
    }


def _append_rows(sheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="E5E7EB")
    section_font = Font(color="111827", bold=True)
    thin_gray = Side(style="thin", color="D1D5DB")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif isinstance(cell.value, str) and cell.column == 1 and cell.value.endswith(":"):
                cell.fill = section_fill
                cell.font = section_font
            if cell.value is not None:
                cell.border = Border(bottom=thin_gray)
            if isinstance(cell.value, (int, float)) and cell.column > 1:
                cell.number_format = '#,##0.0;[Red](#,##0.0);-'
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.font = Font(color="000000")
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 42)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 12)
    sheet.freeze_panes = "A2"


def _write_kv_sheet(sheet, payload: dict[str, Any], *, source_note: str | None = None) -> None:
    sheet.append(["Field", "Value"])
    for key, value in payload.items():
        sheet.append([key, value])
        if source_note:
            sheet.cell(row=sheet.max_row, column=2).comment = Comment(source_note, "Research OS")
    _style_sheet(sheet)


def _build_model_xlsx(
    *,
    ticker: str,
    company_profile: dict[str, Any],
    rows: list[dict[str, Any]],
    ttm_row: dict[str, Any] | None,
    assumptions: dict[str, Any],
    valuation: dict[str, Any],
    sources: dict[str, Any],
    audit: dict[str, Any],
) -> bytes | None:
    if not _model_xlsx_formula_contract_is_ready():
        return None
    value_range = valuation.get("range") or {}
    reliability = valuation.get("reliability") or {}
    if (
        not rows
        or valuation.get("model_version") != "institutional_valuation_v3"
        or not valuation.get("available")
        or valuation.get("status") != "decision_ready"
        or not reliability.get("usable")
        or reliability.get("status") != "high"
        or (valuation.get("price_validation") or {}).get("status") != "validated"
        or (valuation.get("price_validation") or {}).get("usable") is not True
        or audit.get("status") != "pass"
        or _safe_float(value_range.get("central")) is None
    ):
        return None

    workbook = Workbook()
    workbook.remove(workbook.active)

    base_scenario = next((item for item in valuation.get("scenarios", []) if item.get("name") == "base"), {})
    base_scenario_assumptions = base_scenario.get("assumptions") or {}
    current_row = dict(ttm_row or rows[-1])
    capital = valuation.get("cost_of_capital") or {}
    discount_rate = (
        _safe_float(base_scenario_assumptions.get("discount_rate"))
        or _safe_float(base_scenario_assumptions.get("wacc"))
        or _safe_float(base_scenario_assumptions.get("cost_of_equity"))
    )
    terminal_growth = _safe_float(base_scenario_assumptions.get("terminal_growth"))
    latest_shares = _safe_float(current_row.get("diluted_shares")) or 0.0
    latest_cash = _safe_float(current_row.get("cash")) or 0.0
    latest_debt = _safe_float(current_row.get("total_debt")) or 0.0
    latest_revenue = _safe_float(current_row.get("revenue")) or 0.0

    summary_sheet = workbook.create_sheet("Valuation Summary")
    _write_kv_sheet(
        summary_sheet,
        {
            "model_version": valuation.get("model_version"),
            "status": valuation.get("status"),
            "archetype": valuation.get("archetype"),
            "primary_method": valuation.get("primary_method"),
            "cash_flow_basis": valuation.get("cash_flow_basis"),
            "currency": valuation.get("currency"),
            "market_data_as_of": valuation.get("market_data_as_of"),
            "financial_data_as_of": valuation.get("financial_data_as_of"),
            "current_price": valuation.get("current_price"),
            "range_low": value_range.get("low"),
            "range_central": value_range.get("central"),
            "range_high": value_range.get("high"),
            "reliability_status": reliability.get("status"),
            "reliability_score": reliability.get("score"),
            "method_disagreement": reliability.get("method_disagreement"),
            "terminal_value_share": reliability.get("terminal_value_share"),
            "limitations": " | ".join(str(item) for item in reliability.get("limitations", [])),
        },
        source_note="Generated by the canonical institutional valuation contract; status gates determine whether precision may be used for a decision.",
    )

    assumptions_sheet = workbook.create_sheet("Assumptions")
    assumptions_sheet.append(["Field", "Value"])
    _append_rows(
        assumptions_sheet,
        [
            ["Ticker", ticker],
            ["Company", company_profile.get("name")],
            ["Sector", company_profile.get("sector")],
            ["Industry", company_profile.get("industry")],
            ["Model version", valuation.get("model_version")],
            ["Valuation status", valuation.get("status")],
            ["Archetype", valuation.get("archetype")],
            ["Primary method", valuation.get("primary_method")],
            ["Currency", valuation.get("currency")],
            ["Current price", valuation.get("current_price")],
            ["Market data as of", valuation.get("market_data_as_of")],
            ["Financial data as of", valuation.get("financial_data_as_of")],
            ["Diluted shares", latest_shares],
            ["Cash", latest_cash],
            ["Debt", latest_debt],
            ["Starting revenue", latest_revenue],
            ["Discount rate", discount_rate],
            ["Terminal growth", terminal_growth],
            ["Raw beta", capital.get("raw_beta", capital.get("beta"))],
            ["Adjusted beta", capital.get("adjusted_beta", capital.get("beta"))],
            ["Reliability score", reliability.get("score")],
            ["Method disagreement", reliability.get("method_disagreement")],
            ["Terminal value share", reliability.get("terminal_value_share")],
            ["Range low", value_range.get("low")],
            ["Range central", value_range.get("central")],
            ["Range high", value_range.get("high")],
        ],
    )
    for row_number in (18, 19, 22, 23, 24):
        assumptions_sheet.cell(row=row_number, column=2).number_format = "0.0%"
    _style_sheet(assumptions_sheet)

    financials_sheet = workbook.create_sheet("Historical Financials")
    financial_headers = [
        "date",
        "revenue",
        "gross_profit",
        "operating_income",
        "net_income",
        "cash_from_operations",
        "capital_expenditures",
        "free_cash_flow",
        "gross_margin",
        "operating_margin",
        "fcf_margin",
        "roic",
        "diluted_shares",
        "cash",
        "total_debt",
    ]
    financials_sheet.append(financial_headers)
    for row in rows:
        financials_sheet.append([row.get(header) for header in financial_headers])
    if ttm_row:
        financials_sheet.append([ttm_row.get(header) for header in financial_headers])
        financials_sheet.cell(row=financials_sheet.max_row, column=1).comment = Comment(
            "TTM: sum of the latest four discrete fiscal quarters for flow items; latest quarter for balance items; average diluted shares.",
            "Research OS",
        )
    for cell in financials_sheet[1]:
        cell.comment = Comment("Raw statement values are normalized from provider data; calculated columns use formulas in the Python engine.", "Research OS")
    _style_sheet(financials_sheet)

    forecast_sheet = workbook.create_sheet("Forecast")
    forecast_sheet.append(["Year", "Period", "Revenue growth", "Revenue", "Cash-flow margin", "Cash flow", "Discount factor", "Present value"])
    forecast_rows = base_scenario.get("forecast") or []
    for row in forecast_rows:
        year = int(row.get("year") or 0)
        row_idx = forecast_sheet.max_row + 1
        revenue = _safe_float(row.get("revenue")) or 0.0
        cash_flow = _safe_float(row.get("cash_flow")) or 0.0
        growth = _safe_float(row.get("revenue_growth")) or 0.0
        cash_flow_margin = _ratio(cash_flow, revenue) or 0.0
        revenue_formula = f"=Assumptions!$B$17*(1+C{row_idx})" if row_idx == 2 else f"=D{row_idx - 1}*(1+C{row_idx})"
        forecast_sheet.append(
            [
                year,
                row.get("date"),
                growth,
                revenue_formula,
                cash_flow_margin,
                f"=D{row_idx}*E{row_idx}",
                f"=(1+Assumptions!$B$18)^A{row_idx}",
                f"=F{row_idx}/G{row_idx}",
            ]
        )
        forecast_sheet.cell(row=row_idx, column=4).comment = Comment(f"Canonical engine bounded revenue: {revenue}", "Research OS")
        forecast_sheet.cell(row=row_idx, column=6).comment = Comment(f"Canonical engine cash flow: {cash_flow}", "Research OS")
    _style_sheet(forecast_sheet)

    if forecast_rows:
        dcf_sheet = workbook.create_sheet("DCF")
        terminal_row = len(forecast_rows) + 1
        is_fcff = valuation.get("primary_method") == "forward_fcff_dcf"
        _append_rows(
            dcf_sheet,
            [
                ["Line", "Value"],
                ["PV explicit cash flow", f"=SUM(Forecast!H2:H{terminal_row})"],
                ["Terminal cash flow", f"=Forecast!F{terminal_row}*(1+Assumptions!$B$19)"],
                ["Terminal value", "=B3/(Assumptions!$B$18-Assumptions!$B$19)"],
                ["PV terminal value", f"=B4/Forecast!G{terminal_row}"],
                ["Enterprise value", "=B2+B5" if is_fcff else None],
                ["Cash", "=Assumptions!$B$15" if is_fcff else 0.0],
                ["Debt", "=Assumptions!$B$16" if is_fcff else 0.0],
                ["Equity value", "=B6+B7-B8" if is_fcff else "=B2+B5"],
                ["Diluted shares", "=Assumptions!$B$14"],
                ["Intrinsic value per share", "=B9/B10"],
                ["Canonical engine value/share", base_scenario.get("intrinsic_value_per_share")],
                ["Formula reconciliation difference", "=B11-B12"],
            ],
        )
        _style_sheet(dcf_sheet)
    else:
        residual_sheet = workbook.create_sheet("Residual Income")
        residual_sheet.append(["Scenario", "Normalized ROE", "Cost of equity", "Terminal growth", "Terminal value share", "Equity value", "Value/share"])
        for scenario in valuation.get("scenarios", []):
            scenario_assumptions = scenario.get("assumptions") or {}
            residual_sheet.append(
                [
                    scenario.get("name"),
                    scenario_assumptions.get("normalized_roe"),
                    scenario_assumptions.get("cost_of_equity"),
                    scenario_assumptions.get("terminal_growth"),
                    scenario.get("terminal_value_share"),
                    scenario.get("equity_value"),
                    scenario.get("intrinsic_value_per_share"),
                ]
            )
        _style_sheet(residual_sheet)

    reverse_sheet = workbook.create_sheet("Reverse DCF")
    reverse = valuation.get("reverse_dcf") or {}
    _write_kv_sheet(reverse_sheet, {
        "available": reverse.get("available"),
        "status": reverse.get("status"),
        "weight_in_intrinsic_value": reverse.get("weight"),
        "current_price": reverse.get("current_price"),
        "implied_revenue_cagr": reverse.get("implied_revenue_cagr"),
        "bound": reverse.get("bound"),
        "value_at_floor": reverse.get("value_at_floor"),
        "value_at_ceiling": reverse.get("value_at_ceiling"),
        "policy": "Market-implied expectations only; weight is always zero in intrinsic value.",
    })

    multiples_sheet = workbook.create_sheet("Multiples")
    _write_kv_sheet(multiples_sheet, valuation.get("multiples") or {})

    scenarios_sheet = workbook.create_sheet("Scenarios")
    scenarios_sheet.append(["Scenario", "Method", "Discount rate", "Terminal growth", "Terminal value share", "Equity value", "Value/share"])
    for scenario in valuation.get("scenarios", []):
        scenario_assumptions = scenario.get("assumptions") or {}
        scenarios_sheet.append([
            scenario.get("name"),
            scenario.get("method"),
            scenario_assumptions.get("discount_rate") or scenario_assumptions.get("wacc") or scenario_assumptions.get("cost_of_equity"),
            scenario_assumptions.get("terminal_growth"),
            scenario.get("terminal_value_share"),
            scenario.get("equity_value"),
            scenario.get("intrinsic_value_per_share"),
        ])
    _style_sheet(scenarios_sheet)

    if forecast_rows:
        sensitivities_sheet = workbook.create_sheet("Sensitivities")
        growth_grid = [0.01, 0.015, 0.02, 0.025]
        base_rate = discount_rate or 0.10
        rate_grid = [max(0.065, base_rate - 0.015), max(0.065, base_rate - 0.005), min(0.20, base_rate + 0.005), min(0.20, base_rate + 0.015)]
        sensitivities_sheet.append(["Discount rate / terminal growth", *growth_grid])
        terminal_row = len(forecast_rows) + 1
        is_fcff = valuation.get("primary_method") == "forward_fcff_dcf"
        bridge = "+Assumptions!$B$15-Assumptions!$B$16" if is_fcff else ""
        for rate in rate_grid:
            sensitivities_sheet.append([rate, None, None, None, None])
            row_idx = sensitivities_sheet.max_row
            explicit_terms = "+".join(f"Forecast!$F${row}/((1+$A{row_idx})^Forecast!$A${row})" for row in range(2, terminal_row + 1))
            for col_idx in range(2, 6):
                terminal_cell = f"{get_column_letter(col_idx)}$1"
                formula = (
                    f"=({explicit_terms}+(Forecast!$F${terminal_row}*(1+{terminal_cell})/($A{row_idx}-{terminal_cell}))"
                    f"/((1+$A{row_idx})^Forecast!$A${terminal_row}){bridge})/Assumptions!$B$14"
                )
                sensitivities_sheet.cell(row=row_idx, column=col_idx).value = formula
        _style_sheet(sensitivities_sheet)

    sources_sheet = workbook.create_sheet("Sources")
    sources_sheet.append(["source_id", "provider", "endpoint_or_filing", "retrieved_at", "as_of", "status", "row_count", "error"])
    for source in sources.get("records", []):
        sources_sheet.append([
            source.get("source_id"),
            source.get("provider"),
            source.get("endpoint_or_filing"),
            source.get("retrieved_at"),
            source.get("as_of"),
            source.get("status"),
            source.get("row_count"),
            source.get("error"),
        ])
    _style_sheet(sources_sheet)

    audit_sheet = workbook.create_sheet("Audit")
    audit_sheet.append(["severity", "code", "message"])
    for finding in audit.get("findings", []):
        audit_sheet.append([finding.get("severity"), finding.get("code"), finding.get("message")])
    _style_sheet(audit_sheet)

    evidence_sheet = workbook.create_sheet("Evidence Points")
    evidence_sheet.append(["metric", "claim_tag", "source_id", "formula", "normalized_value"])
    for point in sources.get("data_points", []):
        evidence_sheet.append([
            point.get("metric"),
            point.get("claim_tag"),
            point.get("source_id"),
            point.get("formula"),
            point.get("normalized_value"),
        ])
    _style_sheet(evidence_sheet)

    coverage_sheet = workbook.create_sheet("Coverage")
    coverage = sources.get("coverage") or audit.get("coverage") or {}
    coverage_sheet.append(["field", "value"])
    for key, value in coverage.items():
        coverage_sheet.append([key, _json_text(value) if isinstance(value, (list, dict)) else value])
    _style_sheet(coverage_sheet)

    agents_payload = sources.get("agent_outputs") or {}
    agent_claims_sheet = workbook.create_sheet("Agent Claims")
    agent_claims_sheet.append(["agent_id", "agent_name", "claim_tag", "claim", "evidence_refs", "metric_refs"])
    for claim in agents_payload.get("claims", []):
        agent_claims_sheet.append(
            [
                claim.get("agent_id"),
                claim.get("agent_name"),
                claim.get("claim_tag"),
                claim.get("text"),
                ", ".join(str(item) for item in claim.get("evidence_refs", [])),
                ", ".join(str(item) for item in claim.get("metric_refs", [])),
            ]
        )
    _style_sheet(agent_claims_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_downloads(bundle: dict[str, Any]) -> list[dict[str, Any]]:
    ticker = bundle["ticker"]
    downloads = [
        _text_download(f"{ticker}_report.md", "text/markdown", bundle["report_markdown"]),
        _text_download(f"{ticker}_sources.json", "application/json", _json_text(bundle["sources"])),
        _text_download(f"{ticker}_audit.json", "application/json", _json_text(bundle["audit"])),
        _text_download(f"{ticker}_assumptions.yml", "application/yaml", bundle["assumptions_yml"]),
    ]
    if bundle.get("artifacts", {}).get("model_xlsx") is not True:
        return downloads
    model_bytes = _build_model_xlsx(
        ticker=ticker,
        company_profile=bundle["company_profile"],
        rows=bundle["financials"]["annual"],
        ttm_row=bundle["financials"].get("ttm"),
        assumptions=bundle["assumptions"],
        valuation=bundle["valuation"],
        sources=bundle["sources"],
        audit=bundle["audit"],
    )
    if model_bytes:
        downloads.append(
            _bytes_download(
                f"{ticker}_model.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                model_bytes,
            )
        )
    return downloads


def _jsonish_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        nested = value.get("memo_patch")
        if isinstance(nested, str):
            parsed_nested = _jsonish_dict(nested)
            if parsed_nested:
                return {**value, **parsed_nested}
        return value
    text = str(value or "").strip()
    if not text:
        return {}
    fenced = re.search(r"```(?:json)?\s*([\s\S]*?)```", text, flags=re.IGNORECASE)
    candidate = (fenced.group(1) if fenced else text).strip()
    try:
        parsed = json.loads(candidate)
    except json.JSONDecodeError:
        return {"memo_patch": text.replace("```json", "").replace("```", "").strip()}
    return parsed if isinstance(parsed, dict) else {"memo_patch": text}


def _as_text_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value or "").strip()
    return [text] if text else []


def _agent_reader_name(agent: dict[str, Any]) -> str:
    labels = {
        "orchestrator_agent": "Run coordinator",
        "company_profile_agent": "Business profile review",
        "financial_quality_agent": "Financial quality review",
        "valuation_agent": "Valuation review",
        "risk_agent": "Risk review",
        "catalyst_agent": "Filing and catalyst review",
        "red_team_agent": "Red-team challenge",
        "editor_auditor_agent": "Editor and audit gate",
    }
    return labels.get(str(agent.get("id") or ""), str(agent.get("name") or "Analyst review"))


def _agent_status_label(status: Any) -> str:
    value = str(status or "pending").replace("_", " ").strip()
    return value or "pending"


def _agent_reader_lines(agents: list[dict[str, Any]]) -> list[str]:
    if not agents:
        return ["- Analyst desk was not emitted for this run."]
    keep = {
        "orchestrator_agent",
        "financial_quality_agent",
        "valuation_agent",
        "risk_agent",
        "catalyst_agent",
        "red_team_agent",
        "editor_auditor_agent",
    }
    lines = []
    for agent in agents:
        if agent.get("id") not in keep:
            continue
        lines.append(f"- {_agent_reader_name(agent)} ({_agent_status_label(agent.get('status'))}): {agent.get('summary') or 'No summary returned.'}")
    return lines or ["- Analyst desk was not emitted for this run."]


def _final_editor_lines(final_orchestrator: dict[str, Any]) -> list[str]:
    status = final_orchestrator.get("status")
    if status == "ok":
        analysis = _jsonish_dict(final_orchestrator.get("analysis"))
        lines = [
            "## Final editor synthesis",
            "One final editor pass ran after the deterministic engine, source ledger, specialist reviews, and audit. Specialist review roles did not calculate numbers.",
            f"Executive judgment: {analysis.get('executive_judgment') or analysis.get('memo_patch') or 'Completed.'}",
        ]
        strongest = _as_text_list(analysis.get("strongest_points"))
        red_team = _as_text_list(analysis.get("red_team"))
        questions = _as_text_list(analysis.get("open_questions"))
        if strongest:
            lines.extend(["", "What supports the case:", *[f"- {item}" for item in strongest[:4]]])
        if red_team:
            lines.extend(["", "What could break the case:", *[f"- {item}" for item in red_team[:4]]])
        if questions:
            lines.extend(["", "Open checks:", *[f'- {item}' for item in questions[:5]]])
        return lines
    if final_orchestrator.get("enabled"):
        return [
            "## Final editor synthesis",
            f"The one-call final editor was enabled but returned {status or 'no result'}: {final_orchestrator.get('error', 'no synthesis returned')}.",
            "The deterministic analyst desk, valuation, evidence ledger, and audit still remain reproducible.",
        ]
    return [
        "## Final editor synthesis",
        "No final editor synthesis was used in this run. The report below is the deterministic analyst desk output only.",
    ]


def _report_markdown(
    ticker: str,
    profile: dict[str, Any],
    ratios: dict[str, Any],
    valuation: dict[str, Any],
    quality_flags: list[dict[str, Any]],
    audit: dict[str, Any],
    filings: list[dict[str, Any]],
    agent_outputs: dict[str, Any],
) -> str:
    name = profile.get("companyName") or ticker
    sector = profile.get("sector") or "n/a"
    industry = profile.get("industry") or "n/a"
    reverse = valuation.get("reverse_dcf", {})
    implied_growth = reverse.get("implied_revenue_cagr") if reverse.get("available") else None
    value_range = valuation.get("range") or {}
    reliability = valuation.get("reliability") or {}
    price_validation = valuation.get("price_validation") or {}
    backed = bool(
        valuation.get("model_version") == "institutional_valuation_v3"
        and valuation.get("available")
        and valuation.get("status") == "decision_ready"
        and reliability.get("usable") is True
        and reliability.get("status") == "high"
        and price_validation.get("usable") is True
        and price_validation.get("status") == "validated"
        and audit.get("status") == "pass"
    )
    research_range = bool(
        valuation.get("status") == "research_grade"
        and value_range.get("low") is not None
        and value_range.get("high") is not None
    )
    flags = quality_flags or [{"severity": "info", "title": "No accounting quality flags were triggered by the available data."}]
    findings = audit.get("findings") or [{"severity": "info", "message": "No high-severity audit findings."}]
    coverage = audit.get("coverage") or {}
    coverage_line = f"{coverage.get('score', 0)}% ({coverage.get('covered_expected_metrics', 0)}/{coverage.get('expected_metrics', 0)} required metrics)"
    if backed:
        valuation_intro = "The range passed the data, price, method-fit, disagreement, and terminal-value gates required for decision use."
    elif research_range:
        valuation_intro = "A preliminary range is available, but the central estimate is withheld until every reliability gate passes."
    else:
        valuation_intro = "No valuation range is published because the required data or method fit is missing."
    reverse_line = (
        f"{_fmt_pct(implied_growth)} revenue CAGR"
        if implied_growth is not None
        else (reverse.get("bound") or reverse.get("status") or "not available")
    )
    range_line = (
        f"{_fmt_currency(value_range.get('low'))} to {_fmt_currency(value_range.get('high'))}"
        if backed or research_range
        else "withheld until the required data and reliability gates pass"
    )
    statement_authority = coverage.get("statement_authority", "not assessed")
    filing_lines = [
        f"- {item.get('form')} filed {item.get('filing_date')} for period {item.get('report_date') or 'n/a'} ({item.get('accession_number')})"
        for item in filings[:5]
    ] or ["- SEC filings metadata unavailable in this run."]
    agents = agent_outputs.get("agents") or []
    agent_lines = _agent_reader_lines(agents)
    final_orchestrator = agent_outputs.get("final_orchestrator") or {}
    final_orchestrator_lines = _final_editor_lines(final_orchestrator)
    red_team_agent = next((agent for agent in agents if agent.get("id") == "red_team_agent"), {})
    red_team_lines = [
        f"- [{claim.get('claim_tag')}] {claim.get('text')}"
        for claim in red_team_agent.get("claims", [])
    ] or ["- Red-team agent did not emit claims."]

    return "\n".join(
        [
            f"# {ticker} investment memo",
            "",
            "## 1-page investment memo",
            f"- Company: {name}",
            f"- Sector / industry: {sector} / {industry}",
            f"- Evidence state: {audit.get('status')}",
            f"- Evidence coverage: {coverage_line}",
            f"- Latest revenue: {_fmt_currency(ratios.get('latest_revenue'))}",
            f"- Revenue CAGR, 5y: {_fmt_pct(ratios.get('revenue_cagr_5y'))}",
            f"- FCF margin: {_fmt_pct(ratios.get('fcf_margin'))}",
            f"- ROIC: {_fmt_pct(ratios.get('roic'))}",
            f"- Valuation status: {valuation.get('status', 'not_decision_ready')}",
            f"- Valuation method: {valuation.get('primary_method') or 'not available'}",
            f"- Valuation range: {range_line}",
            f"- Central estimate: {_fmt_currency(value_range.get('central')) if backed else 'withheld until all reliability gates pass'}",
            f"- Reliability: {reliability.get('status', 'blocked')} ({_fmt_pct(reliability.get('score'))})",
            f"- Market data date: {valuation.get('market_data_as_of') or 'n/a'}",
            f"- Financial data date: {valuation.get('financial_data_as_of') or 'n/a'}",
            f"- Reverse DCF expectation: {reverse_line}",
            "",
            "## Valuation suite",
            valuation_intro,
            f"Primary method: {valuation.get('primary_method') or 'not available'}; company archetype: {valuation.get('archetype') or 'not available'}.",
            *[f"Limitation: {item}" for item in reliability.get("limitations", [])],
            f"Statement authority: {statement_authority}. SEC EDGAR is used for filing metadata unless XBRL fact ingestion is explicitly present in sources.json.",
            "",
            "## Analyst desk",
            "How to read this: Python pulls the data and calculates the metrics. The analyst desk is a set of reproducible review roles that read those audited outputs, challenge the case, and point to open checks.",
            *agent_lines,
            "",
            *final_orchestrator_lines,
            "",
            "## Red-team memo",
            *red_team_lines,
            "",
            "## Accounting quality flags",
            *[f"- [{item.get('severity')}] {item.get('title')}" for item in flags],
            "",
            "## Authoritative filings",
            *filing_lines,
            "",
            "## Audit findings",
            *[f"- [{item.get('severity')}] {item.get('message')}" for item in findings],
            "",
            "## Source appendix",
            f"See sources.json for provider endpoints, timestamps, row counts, errors, and coverage gaps. Statement authority: {statement_authority}.",
            "",
        ]
    )


def build_equity_research_bundle(
    ticker: str,
    *,
    mode: str = "quick",
    paths: PathConfigLike | None = None,
    fmp_client: FMPClient | None = None,
    sec_client: SECEdgarClient | None = None,
    yfinance_factory: Any | None = None,
    enable_yfinance: bool = False,
    llm_client: Any | None = None,
    enable_llm: bool | None = None,
) -> dict[str, Any]:
    symbol = clean_ticker(ticker)
    if not symbol:
        return {
            "ok": False,
            "error": "Ticker is required.",
            "audit": {
                "status": "needs_attention",
                "findings": [{"severity": "high", "code": "missing_ticker", "message": "Ticker is required."}],
            },
        }

    if fmp_client is None and paths is not None:
        fmp_client = FMPClient.from_env(paths.cache_root)
    if sec_client is None and paths is not None:
        sec_client = SECEdgarClient.from_env(paths.cache_root)

    statement_source_ids = dict(DEFAULT_STATEMENT_SOURCE_IDS)
    quarterly_statement_source_ids = dict(QUARTERLY_STATEMENT_SOURCE_IDS)
    profile, frames, sources = _load_fmp_payloads(symbol, paths, fmp_client)
    use_yfinance = yfinance_factory is not None or enable_yfinance is True
    if use_yfinance:
        include_yfinance_statements = any(
            frames.get(key, pd.DataFrame()).empty
            for key in (
                "income",
                "cash_flow",
                "balance",
                "income_quarterly",
                "cash_flow_quarterly",
                "balance_quarterly",
            )
        )
        yfinance_profile, yfinance_frames, yfinance_sources = _load_yfinance_payloads(
            symbol,
            yfinance_factory,
            include_statements=include_yfinance_statements,
        )
        sources.extend(yfinance_sources)
        if yfinance_profile:
            if not profile:
                profile = yfinance_profile
            else:
                profile = {
                    key: profile.get(key) if _has_value(profile.get(key)) else value
                    for key, value in {**yfinance_profile, **profile}.items()
                }
                profile["source_family"] = "FMP"
        yfinance_quote = yfinance_frames.get("quote") or {}
        if not frames.get("quote") and yfinance_quote:
            frames["quote"] = yfinance_quote
        yfinance_prices = yfinance_frames.get("prices", pd.DataFrame())
        if isinstance(yfinance_prices, pd.DataFrame) and not yfinance_prices.empty:
            # A Yahoo close provides either a strict same-provider pair or an
            # independent check on an FMP quote. Keep the provider identity on
            # each observation; do not blend prices into a synthetic series.
            frames["prices"] = yfinance_prices
        if not frames.get("shares_float") and yfinance_frames.get("shares_float"):
            frames["shares_float"] = yfinance_frames["shares_float"]
    else:
        yfinance_frames = {}
    filings, sec_sources = _load_sec_filings(symbol, sec_client)
    sources.extend(sec_sources)
    sec_frames, sec_fact_sources, sec_company_facts = _load_sec_company_facts(symbol, sec_client)
    sources.extend(sec_fact_sources)
    if any(frames.get(key, pd.DataFrame()).empty for key in DEFAULT_STATEMENT_SOURCE_IDS):
        for statement_key, source_id in SEC_STATEMENT_SOURCE_IDS.items():
            if frames.get(statement_key, pd.DataFrame()).empty and not sec_frames.get(statement_key, pd.DataFrame()).empty:
                frames[statement_key] = sec_frames[statement_key]
                statement_source_ids[statement_key] = source_id
    for statement_key, source_id in YFINANCE_STATEMENT_SOURCE_IDS.items():
        yfinance_frame = yfinance_frames.get(statement_key, pd.DataFrame())
        if frames.get(statement_key, pd.DataFrame()).empty and isinstance(yfinance_frame, pd.DataFrame) and not yfinance_frame.empty:
            frames[statement_key] = yfinance_frame
            statement_source_ids[statement_key] = source_id
    for statement_key, source_id in YFINANCE_QUARTERLY_STATEMENT_SOURCE_IDS.items():
        frame_key = f"{statement_key}_quarterly"
        yfinance_frame = yfinance_frames.get(frame_key, pd.DataFrame())
        if frames.get(frame_key, pd.DataFrame()).empty and isinstance(yfinance_frame, pd.DataFrame) and not yfinance_frame.empty:
            frames[frame_key] = yfinance_frame
            quarterly_statement_source_ids[statement_key] = source_id
    balance_enrichments = _enrich_balance_frames_with_sec(frames, sec_frames)
    if balance_enrichments:
        balance_source = next(
            (source for source in sources if source.get("source_id") == SEC_STATEMENT_SOURCE_IDS["balance"]),
            None,
        )
        if balance_source is not None:
            balance_source["field_enrichments"] = balance_enrichments
            balance_source["enriched_field_count"] = len(balance_enrichments)
    preliminary_assumptions = {"normalized_tax_rate": DEFAULT_TAX_RATE}
    rows = _normalize_financials(frames, preliminary_assumptions["normalized_tax_rate"])
    preliminary_ttm = _build_ttm_row(
        frames,
        preliminary_assumptions["normalized_tax_rate"],
        sec_company_facts=sec_company_facts,
    )
    assumptions = _derive_assumptions(rows, preliminary_ttm)
    rows = _normalize_financials(frames, assumptions["normalized_tax_rate"])
    ttm_row = _build_ttm_row(
        frames,
        assumptions["normalized_tax_rate"],
        sec_company_facts=sec_company_facts,
    )
    assumptions = _derive_assumptions(rows, ttm_row)
    ratios = _build_ratios(rows, ttm_row)
    quality_flags = _quality_flags(rows)
    share_denominator_reconciliation = _reconcile_current_share_count(
        symbol,
        ttm_row,
        frames.get("shares_float") if isinstance(frames.get("shares_float"), dict) else {},
    )
    if ttm_row:
        ttm_row.setdefault("ttm_validation", {})["share_denominator_reconciliation"] = share_denominator_reconciliation
    valuation = _build_valuation(ticker, rows, ttm_row, profile, frames, assumptions, sources)
    valuation = _apply_current_share_count_gate(valuation, share_denominator_reconciliation)
    independent_observation = (valuation.get("price_validation") or {}).get("independent_observation") or {}
    if independent_observation:
        observation_source_id = str(independent_observation.get("source_id") or "market:independent-close")
        existing_observation_source = next(
            (source for source in sources if source.get("source_id") == observation_source_id),
            None,
        )
        if existing_observation_source is not None:
            existing_observation_source.update(
                {
                    "as_of": independent_observation.get("as_of") or existing_observation_source.get("as_of"),
                    "observed_price": independent_observation.get("price"),
                    "currency": independent_observation.get("currency"),
                    "source_family": independent_observation.get("source_family"),
                }
            )
        else:
            sources.append(
                _source_record(
                    observation_source_id,
                    str(independent_observation.get("source_family") or "independent market"),
                    "independent market close observation",
                    "ok",
                    as_of=independent_observation.get("as_of"),
                    observed_price=independent_observation.get("price"),
                    currency=independent_observation.get("currency"),
                    source_family=independent_observation.get("source_family"),
                )
            )
    using_sec_as_primary = _all_statement_families_use_sec(statement_source_ids)
    country = str(profile.get("country") or "").upper().strip()
    sec_reconciliation_required = country in {"US", "USA", "UNITED STATES", "UNITED STATES OF AMERICA"}
    if using_sec_as_primary:
        reconciliation = {"status": "primary_sec", "passed": True, "pass_ratio": 1.0, "checks": []}
    else:
        sec_rows = _normalize_financials(sec_frames, assumptions["normalized_tax_rate"])
        reconciliation = _reconcile_statement_rows(rows, sec_rows) if sec_rows else {
            "status": "insufficient_overlap",
            "passed": False,
            "pass_ratio": None,
            "checks": [],
            "overlap_periods": [],
        }
        if not sec_reconciliation_required and reconciliation.get("status") == "insufficient_overlap":
            reconciliation = {**reconciliation, "status": "not_required", "passed": True}
    sources.append(
        _source_record(
            SEC_RECONCILIATION_SOURCE_ID,
            "deterministic-audit",
            "Normalized statements vs SEC Company Facts/XBRL",
            "ok" if reconciliation.get("passed") else ("error" if reconciliation.get("status") == "mismatch" else "unavailable"),
            reconciliation_status=reconciliation.get("status"),
            pass_ratio=reconciliation.get("pass_ratio"),
            row_count=len(reconciliation.get("checks") or []),
            overlap_periods=reconciliation.get("overlap_periods") or [],
            checks=reconciliation.get("checks") or [],
        )
    )
    valuation = _apply_statement_reconciliation_gate(
        valuation,
        reconciliation,
        required=sec_reconciliation_required,
    )
    model_base_case = next((item for item in valuation.get("scenarios", []) if item.get("name") == "base"), {})
    model_base_assumptions = model_base_case.get("assumptions") or {}
    market_requirements = valuation.get("market_requirements") or {}
    model_discount_rate = (
        model_base_assumptions.get("discount_rate")
        or model_base_assumptions.get("wacc")
        or model_base_assumptions.get("cost_of_equity")
        or market_requirements.get("discount_rate")
    )
    model_terminal_growth = (
        model_base_assumptions.get("terminal_growth")
        if model_base_assumptions.get("terminal_growth") is not None
        else market_requirements.get("terminal_growth")
    )

    if ttm_row:
        latest_revenue_point = _data_point(
            "latest_revenue",
            ratios.get("latest_revenue"),
            "calculated_metric",
            formula="sum of four validated quarterly revenue observations; see financials.ttm.revenue lineage",
        )
        latest_shares_point = _data_point(
            "latest_diluted_shares",
            ratios.get("latest_diluted_shares"),
            "calculated_metric",
            formula="average diluted shares across four validated quarters; see financials.ttm.diluted_shares lineage",
        )
    else:
        latest_revenue_point = _data_point("latest_revenue", ratios.get("latest_revenue"), "sourced_fact", statement_source_ids.get("income"))
        latest_shares_point = _data_point("latest_diluted_shares", ratios.get("latest_diluted_shares"), "sourced_fact", statement_source_ids.get("income"))

    profile_source_id = (
        "yfinance:profile"
        if str(profile.get("source_family") or "").lower() in {"yfinance", "yahoo finance", "yahoo"}
        else "fmp:profile"
    )
    quote_provider_family = str((valuation.get("price_validation") or {}).get("provider_family") or "").lower()
    quote_record_source_id = (
        "yfinance:quote"
        if quote_provider_family in {"yfinance", "yahoo finance", "yahoo"}
        else "fmp:quote"
    )
    active_quote = frames.get("quote") if isinstance(frames.get("quote"), dict) else {}
    price_frame = frames.get("prices", pd.DataFrame())
    price_frame_source_family = ""
    if isinstance(price_frame, pd.DataFrame) and not price_frame.empty:
        latest_price_rows = price_frame.copy()
        if "date" in latest_price_rows.columns:
            latest_price_rows = latest_price_rows.sort_values("date")
        price_frame_source_family = str(latest_price_rows.iloc[-1].get("source_family") or "FMP").lower()
    market_price_source_id = (
        quote_record_source_id
        if _safe_float(active_quote.get("price")) is not None
        else "yfinance:prices"
        if price_frame_source_family in {"yfinance", "yahoo finance", "yahoo"}
        else "fmp:prices"
        if isinstance(price_frame, pd.DataFrame) and not price_frame.empty
        else profile_source_id
    )
    market_cap_source_id = (
        quote_record_source_id
        if _safe_float(active_quote.get("marketCap")) is not None or _safe_float(active_quote.get("market_cap")) is not None
        else profile_source_id
    )
    shares_float_source_id = (
        "yfinance:shares-float"
        if str((frames.get("shares_float") or {}).get("source_family") or "").lower() in {"yfinance", "yahoo finance", "yahoo"}
        else "fmp:shares-float"
    )
    active_screening_statement_sources = (
        quarterly_statement_source_ids if ttm_row else statement_source_ids
    )

    data_points = [
        _data_point("company_profile", profile.get("companyName") or symbol, "sourced_fact", profile_source_id),
        latest_revenue_point,
        latest_shares_point,
        _data_point("latest_free_cash_flow", ratios.get("latest_fcf"), "calculated_metric", formula="cash_from_operations - abs(capital_expenditures)"),
        _data_point("revenue_cagr_5y", ratios.get("revenue_cagr_5y"), "calculated_metric", formula="(Revenue_t / Revenue_0) ** (1 / years) - 1"),
        _data_point("gross_margin", ratios.get("gross_margin"), "calculated_metric", formula="gross_profit / revenue"),
        _data_point("operating_margin", ratios.get("operating_margin"), "calculated_metric", formula="operating_income / revenue"),
        _data_point("fcf_margin", ratios.get("fcf_margin"), "calculated_metric", formula="free_cash_flow / revenue"),
        _data_point("roic", ratios.get("roic"), "calculated_metric", formula="NOPAT / average_invested_capital"),
        _data_point("net_debt", ratios.get("net_debt"), "calculated_metric", formula="total_debt - cash_and_equivalents"),
        _data_point("base_fcf_margin", assumptions.get("base_fcf_margin"), "assumption", formula="historical cash-flow reference; cyclical valuation separately includes weak and loss years"),
        _data_point("wacc", model_discount_rate, "assumption", formula="price-independent operating risk rate; no debt tax shield is included without an explicit reproducible debt schedule"),
        _data_point("terminal_growth", model_terminal_growth, "assumption", formula="bounded by company archetype and below the discount rate"),
    ]
    current_basic_shares = _safe_float(share_denominator_reconciliation.get("current_basic_outstanding_shares"))
    if current_basic_shares is not None:
        data_points.extend(
            [
                _data_point(
                    "current_basic_outstanding_shares",
                    current_basic_shares,
                    "sourced_fact",
                    shares_float_source_id,
                ),
                _data_point(
                    "current_share_count_relative_difference",
                    share_denominator_reconciliation.get("relative_difference"),
                    "calculated_metric",
                    formula="abs(current basic outstanding shares - TTM diluted weighted-average shares) / TTM diluted weighted-average shares",
                ),
            ]
        )
    data_points.extend(_financial_data_points(rows, statement_source_ids))
    data_points.extend(_ttm_data_points(ttm_row, quarterly_statement_source_ids))
    if _has_value(valuation.get("current_price")):
        data_points.append(
            _data_point("current_price", valuation.get("current_price"), "sourced_fact", market_price_source_id)
        )
    screening = valuation.get("screening_analysis") or {}
    observed_screening = screening.get("observed") or {}
    if screening.get("available") is True:
        screening_sources = {
            "current_price": market_price_source_id,
            "market_cap": market_cap_source_id,
            "revenue": active_screening_statement_sources.get("income"),
            "free_cash_flow": active_screening_statement_sources.get("cash_flow"),
            "cash": active_screening_statement_sources.get("balance"),
            "total_debt": active_screening_statement_sources.get("balance"),
            "diluted_shares": active_screening_statement_sources.get("income"),
        }
        for field, source_id in screening_sources.items():
            value = observed_screening.get(field)
            if _has_value(value) and source_id:
                data_points.append(
                    _data_point(f"screening.{field}", value, "sourced_fact", source_id)
                )
    if valuation.get("available"):
        data_points.extend(
            [
                _data_point("reverse_dcf_status", valuation.get("reverse_dcf", {}).get("status"), "calculated_metric", formula="bounded reverse-valuation solve status; reverse DCF has zero intrinsic-value weight"),
                _data_point("valuation_reverse_dcf_implied_revenue_cagr", valuation.get("reverse_dcf", {}).get("implied_revenue_cagr"), "calculated_metric", formula="binary search for growth where DCF value equals current price"),
                _data_point("ev_to_sales", valuation.get("multiples", {}).get("ev_to_sales"), "calculated_metric", formula="enterprise_value / latest_revenue"),
                _data_point("price_to_fcf", valuation.get("multiples", {}).get("price_to_fcf"), "calculated_metric", formula="market_cap / latest_free_cash_flow"),
            ]
        )
        data_points.extend(_valuation_data_points(valuation))
    elif market_requirements.get("available") is True:
        market_assets_point = _data_point(
            "market_requirement_assets_added",
            market_requirements.get("assets_added"),
            "calculated_metric",
            formula="excess cash plus explicitly sourced non-operating investments added to enterprise value",
        )
        market_assets_point["source_ids"] = [
            QUARTERLY_STATEMENT_SOURCE_IDS["balance"],
            TTM_STATEMENT_SOURCE_IDS["balance"],
        ]
        market_obligations_point = _data_point(
            "market_requirement_obligations_deducted",
            market_requirements.get("obligations_deducted"),
            "calculated_metric",
            formula="debt plus explicitly sourced common-equity senior claims deducted from enterprise value",
        )
        market_obligations_point["source_ids"] = [
            QUARTERLY_STATEMENT_SOURCE_IDS["balance"],
            TTM_STATEMENT_SOURCE_IDS["balance"],
        ]
        data_points.extend(
            [
                _data_point(
                    "reverse_dcf_status",
                    market_requirements.get("status"),
                    "calculated_metric",
                    formula="bounded reverse-valuation solve status; no intrinsic value is published",
                ),
                _data_point(
                    "market_requirement_implied_revenue_cagr",
                    market_requirements.get("implied_revenue_cagr"),
                    "calculated_metric",
                    formula="binary search for growth where price-implied enterprise value equals discounted operating cash flow",
                ),
                _data_point(
                    "market_requirement_normalized_margin",
                    market_requirements.get("normalized_margin"),
                    "calculated_metric",
                    formula="through-cycle operating FCFF after SBC divided by revenue, using a complete observed cycle",
                ),
                market_assets_point,
                market_obligations_point,
            ]
        )
    if filings:
        data_points.append(_data_point("latest_sec_filing", filings[0].get("accession_number"), "sourced_fact", "sec:submissions"))
    audit = _audit_bundle(symbol, rows, sources, valuation, data_points)
    coverage = audit.get("coverage", {})
    checklist = _checklist_scores(ratios, quality_flags, valuation, audit)
    valuation_backed = _valuation_precision_is_backed(valuation, audit)
    model_xlsx_ready = valuation_backed and _model_xlsx_formula_contract_is_ready()
    output_valuation = _valuation_for_output(valuation, audit)
    output_data_points = _data_points_for_output(data_points, backed=valuation_backed)
    output_assumptions = _canonical_assumptions(valuation, assumptions)

    company_profile = {
        "name": profile.get("companyName") or symbol,
        "sector": profile.get("sector"),
        "industry": profile.get("industry"),
        "country": profile.get("country"),
        "currency": profile.get("currency"),
        "exchange": profile.get("exchangeShortName") or profile.get("exchange"),
        "beta": _safe_float(profile.get("beta")),
        "market_cap": _safe_float(profile.get("marketCap")) or _safe_float(profile.get("mktCap")),
        "description": profile.get("description") if mode == "full" else None,
    }
    agent_outputs = build_agent_outputs(
        ticker=symbol,
        profile=profile,
        rows=rows,
        ratios=ratios,
        valuation=output_valuation,
        quality_flags=quality_flags,
        audit=audit,
        sources=sources,
        filings=filings,
    )
    agent_outputs["final_orchestrator"] = run_final_orchestrator_llm(
        ticker=symbol,
        profile=profile,
        rows=rows,
        ratios=ratios,
        valuation=output_valuation,
        quality_flags=quality_flags,
        audit=audit,
        agent_outputs=agent_outputs,
        filings=filings,
        llm_client=llm_client,
        enabled=enable_llm,
    )
    bundle = {
        "ok": True,
        "ticker": symbol,
        "mode": "full" if mode == "full" else "quick",
        "generated_at": _now_iso(),
        "company_profile": company_profile,
        "financials": {
            "annual": rows,
            "ttm": ttm_row,
            "ratios": ratios,
            "quality_flags": quality_flags,
        },
        "filings": {
            "recent": filings,
        },
        "valuation": output_valuation,
        "checklist_score": checklist,
        "agents": agent_outputs,
        "report_markdown": _report_markdown(symbol, profile, ratios, output_valuation, quality_flags, audit, filings, agent_outputs),
        "sources": {
            "records": sources,
            "data_points": output_data_points,
            "coverage": coverage,
            "statement_source_ids": statement_source_ids,
            "claims": agent_outputs.get("claims", []),
            "agent_outputs": agent_outputs,
        },
        "audit": audit,
        "assumptions": output_assumptions,
        "assumptions_yml": _assumptions_yml(output_assumptions),
        "artifacts": {
            "report_md": True,
            "model_xlsx": model_xlsx_ready,
            "model_xlsx_reason": None if model_xlsx_ready else "withheld_until_headline_formulas_reconcile_to_zero",
            "sources_json": True,
            "audit_json": True,
            "assumptions_yml": True,
            "note": "Artifacts reproduce the routed valuation method, dates, scenario range, price checks, and reliability gates.",
        },
    }
    bundle["downloads"] = _build_downloads(bundle)
    return bundle
